import sqlite3
import pandas as pd
import numpy as np
from pathlib import Path
from typing import Dict, List, Tuple, Optional
import logging


class ExcelResultsCache:
    """Optimoitu Excel-tulosten generointi staging-tietokannan avulla"""

    def __init__(
        self,
        analysis_db_path: str = "analysis/analysis.db",
        osake_db_path: str = "data/osakedata.db",
        results_db_path: str = "data/results.db",
    ):

        self.analysis_db = analysis_db_path
        self.osake_db = osake_db_path
        self.results_db = results_db_path

        # Luo results.db ja tarvittavat taulut
        self._init_results_database()

    def _init_results_database(self):
        """Alusta results.db tietokanta"""
        with sqlite3.connect(self.results_db) as conn:
            conn.executescript(
                """
                -- Staging-taulu Excel-tuloksia varten (78 saraketta SARAKKEET_DOKUMENTAATIO.md mukaan)
                CREATE TABLE IF NOT EXISTS excel_staging (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    
                    -- Perustiedot (1-3)
                    osake TEXT NOT NULL,               -- 1. osake (ticker-symboli)
                    pvm TEXT NOT NULL,                 -- 2. pvm (YYYY-MM-DD)
                    kynttila INTEGER NOT NULL,         -- 3. kynttila (1-7)
                    
                    -- T-1 Edellinen päivä (4-7)
                    t_1_alin REAL,                     -- 4. t_1_alin (normalisoitu)
                    t_1_ylin REAL,                     -- 5. t_1_ylin (normalisoitu)
                    t_1_bodi REAL,                     -- 6. t_1_bodi (% avaus)
                    t_1_bodi_colour INTEGER,           -- 7. t_1_bodi_colour (0/1)
                    
                    -- T0 Käännekynttilä (8-11)
                    t0_alin REAL,                      -- 8. t0_alin (aina 100.00)
                    t0_ylin REAL,                      -- 9. t0_ylin (normalisoitu)
                    t0_bodi REAL,                      -- 10. t0_bodi (% avaus)
                    t0_bodi_colour INTEGER,            -- 11. t0_bodi_colour (0/1)
                    
                    -- T1 Seuraava päivä (12-15)
                    t1_alin REAL,                      -- 12. t1_alin (normalisoitu)
                    t1_ylin REAL,                      -- 13. t1_ylin (normalisoitu)
                    t1_bodi REAL,                      -- 14. t1_bodi (% avaus)
                    t1_bodi_colour INTEGER,            -- 15. t1_bodi_colour (0/1)
                    
                    -- Historialliset hinnat (16-18)
                    t_5 REAL,                          -- 16. t_5 (5 pv sitten)
                    t_10 REAL,                         -- 17. t_10 (10 pv sitten)
                    t_20 REAL,                         -- 18. t_20 (20 pv sitten)
                    
                    -- Tulevat hinnat (19-24)
                    t5 REAL,                           -- 19. t5 (5 pv eteenpäin)
                    t10 REAL,                          -- 20. t10 (10 pv eteenpäin)
                    t20 REAL,                          -- 21. t20 (20 pv eteenpäin)
                    t40 REAL,                          -- 22. t40 (40 pv eteenpäin)
                    t60 REAL,                          -- 23. t60 (60 pv eteenpäin)
                    t252 REAL,                         -- 24. t252 (252 pv eteenpäin)
                    
                    -- Volatiliteetti (25-29)
                    vol_5 REAL,                        -- 25. vol_5 (5 pv volatiliteetti)
                    vol_10 REAL,                       -- 26. vol_10 (10 pv volatiliteetti)
                    vol_20 REAL,                       -- 27. vol_20 (20 pv volatiliteetti)
                    vol_60 REAL,                       -- 28. vol_60 (60 pv volatiliteetti)
                    vol_252 REAL,                      -- 29. vol_252 (252 pv volatiliteetti)
                    
                    -- Volyymit (30-39)
                    vol_avg_10 REAL,                   -- 30. vol_avg_10 (10 pv keskivol)
                    vol_avg_20 REAL,                   -- 31. vol_avg_20 (20 pv keskivol)
                    vol_avg_60 REAL,                   -- 32. vol_avg_60 (60 pv keskivol)
                    vol_t_1_vs_avg10 REAL,             -- 33. vol_t_1_vs_avg10
                    vol_t0_vs_avg10 REAL,              -- 34. vol_t0_vs_avg10
                    vol_t1_vs_avg10 REAL,              -- 35. vol_t1_vs_avg10
                    vol_t_1_vs_avg20 REAL,             -- 36. vol_t_1_vs_avg20
                    vol_t0_vs_avg20 REAL,              -- 37. vol_t0_vs_avg20
                    vol_t1_vs_avg20 REAL,              -- 38. vol_t1_vs_avg20
                    vol_spike INTEGER,                 -- 39. vol_spike (0/1)
                    
                    -- Liukuvat keskiarvot (40-56)
                    ma_5 REAL,                         -- 40. ma_5 (5 pv MA)
                    ma_10 REAL,                        -- 41. ma_10 (10 pv MA)
                    ma_20 REAL,                        -- 42. ma_20 (20 pv MA)
                    ma_50 REAL,                        -- 43. ma_50 (50 pv MA)
                    ma_100 REAL,                       -- 44. ma_100 (100 pv MA)
                    ma_200 REAL,                       -- 45. ma_200 (200 pv MA)
                    dist_ma_5 REAL,                    -- 46. dist_ma_5 (etäisyys %)
                    dist_ma_10 REAL,                   -- 47. dist_ma_10 (etäisyys %)
                    dist_ma_20 REAL,                   -- 48. dist_ma_20 (etäisyys %)
                    dist_ma_50 REAL,                   -- 49. dist_ma_50 (etäisyys %)
                    dist_ma_100 REAL,                  -- 50. dist_ma_100 (etäisyys %)
                    dist_ma_200 REAL,                  -- 51. dist_ma_200 (etäisyys %)
                    ma_5_slope REAL,                   -- 52. ma_5_slope (kulmakerroin)
                    ma_10_slope REAL,                  -- 53. ma_10_slope (kulmakerroin)
                    ma_20_slope REAL,                  -- 54. ma_20_slope (kulmakerroin)
                    ma_50_slope REAL,                  -- 55. ma_50_slope (kulmakerroin)
                    ma_200_slope REAL,                 -- 56. ma_200_slope (kulmakerroin)
                    
                    -- S&P 500 tiedot (57-67)
                    sp500_t_1 REAL,                    -- 57. sp500_t_1
                    sp500_t0 REAL,                     -- 58. sp500_t0
                    sp500_t1 REAL,                     -- 59. sp500_t1
                    sp500_ma_20 REAL,                  -- 60. sp500_ma_20
                    sp500_ma_50 REAL,                  -- 61. sp500_ma_50
                    sp500_ma_200 REAL,                 -- 62. sp500_ma_200
                    sp500_change_t0 REAL,              -- 63. sp500_change_t0 (%)
                    sp500_change_t1 REAL,              -- 64. sp500_change_t1 (%)
                    sp500_vs_ma20 REAL,                -- 65. sp500_vs_ma20 (%)
                    sp500_vs_ma50 REAL,                -- 66. sp500_vs_ma50 (%)
                    sp500_vs_ma200 REAL,               -- 67. sp500_vs_ma200 (%)
                    
                    -- Nasdaq 100 tiedot (68-78)
                    nasdaq_t_1 REAL,                   -- 68. nasdaq_t_1
                    nasdaq_t0 REAL,                    -- 69. nasdaq_t0
                    nasdaq_t1 REAL,                    -- 70. nasdaq_t1
                    nasdaq_ma_20 REAL,                 -- 71. nasdaq_ma_20
                    nasdaq_ma_50 REAL,                 -- 72. nasdaq_ma_50
                    nasdaq_ma_200 REAL,                -- 73. nasdaq_ma_200
                    nasdaq_change_t0 REAL,             -- 74. nasdaq_change_t0 (%)
                    nasdaq_change_t1 REAL,             -- 75. nasdaq_change_t1 (%)
                    nasdaq_vs_ma20 REAL,               -- 76. nasdaq_vs_ma20 (%)
                    nasdaq_vs_ma50 REAL,               -- 77. nasdaq_vs_ma50 (%)
                    nasdaq_vs_ma200 REAL,              -- 78. nasdaq_vs_ma200 (%)
                    
                    -- Metadata
                    calculated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    
                    UNIQUE(osake, pvm, kynttila)
                );
                
                -- Indeksit nopeaa hakua varten
                CREATE INDEX IF NOT EXISTS idx_excel_staging_osake ON excel_staging(osake);
                CREATE INDEX IF NOT EXISTS idx_excel_staging_pvm ON excel_staging(pvm);
                CREATE INDEX IF NOT EXISTS idx_excel_staging_osake_pvm ON excel_staging(osake, pvm);
                
                -- Cache-metatieto
                CREATE TABLE IF NOT EXISTS cache_metadata (
                    key TEXT PRIMARY KEY,
                    value TEXT,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                );
            """
            )

    def is_cache_fresh(self) -> bool:
        """Tarkista onko cache tuoreempi kuin lähdedata"""
        try:
            with sqlite3.connect(self.results_db) as conn:
                cache_ts = conn.execute(
                    """
                    SELECT value FROM cache_metadata 
                    WHERE key = 'last_full_rebuild'
                """
                ).fetchone()

            if not cache_ts:
                return False

            # Tarkista analysis.db:n viimeinen muutos
            with sqlite3.connect(self.analysis_db) as conn:
                source_count = conn.execute(
                    """
                    SELECT COUNT(*) FROM analysis_findings
                """
                ).fetchone()[0]

            with sqlite3.connect(self.results_db) as conn:
                cached_count = conn.execute(
                    """
                    SELECT value FROM cache_metadata 
                    WHERE key = 'source_count'
                """
                ).fetchone()

            if not cached_count or int(cached_count[0]) != source_count:
                return False

            return True

        except Exception as e:
            logging.warning(f"Cache freshness check failed: {e}")
            return False

    def rebuild_staging_optimized(
        self, progress_callback=None, limit_rows: int = None, ticker_filter: str = None
    ):
        """Rakenna staging-taulu optimoidusti SQL:llä"""

        def update_progress(step: str, current: int, total: int):
            if progress_callback:
                progress_callback(step, current, total)

        update_progress("Alustetaan staging-taulua...", 0, 100)

        # Tyhjennä vanha staging
        with sqlite3.connect(self.results_db) as conn:
            conn.execute("DELETE FROM excel_staging")
            conn.execute("DELETE FROM cache_metadata")

        update_progress("Haetaan löydökset...", 10, 100)

        # Hae löydökset analysis.db:stä (ticker-filtterillä jos annettu)
        with sqlite3.connect(self.analysis_db) as analysis_conn:
            if ticker_filter:
                print(f"🎯 Suodatetaan ticker: {ticker_filter}")
                findings = analysis_conn.execute(
                    """
                    SELECT ticker, date, candle 
                    FROM analysis_findings 
                    WHERE candle IS NOT NULL AND candle != '' AND ticker = ?
                    ORDER BY ticker, date
                """,
                    (ticker_filter,),
                ).fetchall()
            else:
                print("🌐 Haetaan kaikki tickerit")
                findings = analysis_conn.execute(
                    """
                    SELECT ticker, date, candle 
                    FROM analysis_findings 
                    WHERE candle IS NOT NULL AND candle != ''
                    ORDER BY ticker, date
                """
                ).fetchall()

        if not findings:
            logging.warning("Ei löydöksiä analysis.db:ssä")
            return

        update_progress("Ryhmitellään osakkeittain...", 20, 100)

        # Ryhmittele osakkeittain
        by_ticker = {}
        for ticker, date, candle in findings:
            if ticker not in by_ticker:
                by_ticker[ticker] = []
            by_ticker[ticker].append((date, candle))

        total_tickers = len(by_ticker)
        processed_tickers = 0

        # Rajoita prosessoitavien osakkeiden määrää limit_rows:n perusteella
        if limit_rows:
            ticker_limit = min(limit_rows, total_tickers)
            by_ticker = dict(list(by_ticker.items())[:ticker_limit])
            total_tickers = len(by_ticker)  # Prosessoi osake kerrallaan
        with sqlite3.connect(self.osake_db) as osake_conn:
            with sqlite3.connect(self.results_db) as results_conn:

                for ticker, ticker_findings in by_ticker.items():
                    self._process_ticker_optimized(
                        ticker, ticker_findings, osake_conn, results_conn
                    )

                    processed_tickers += 1
                    progress_pct = 20 + int((processed_tickers / total_tickers) * 70)
                    update_progress(f"Prosessoidaan {ticker}...", progress_pct, 100)

        # Päivitä cache-metadata
        with sqlite3.connect(self.results_db) as conn:
            source_count = len(findings)
            conn.execute(
                """
                INSERT OR REPLACE INTO cache_metadata (key, value) VALUES 
                ('last_full_rebuild', datetime('now')),
                ('source_count', ?)
            """,
                (source_count,),
            )

        update_progress("Valmis!", 100, 100)

    def _process_ticker_optimized(
        self, ticker: str, findings: List[Tuple[str, str]], osake_conn, results_conn
    ):
        """Prosessoi yhden osakkeen kaikki löydökset optimoidusti"""

        # Hae kaikki osakkeen hinnat kerralla
        try:
            df = pd.read_sql(
                """
                SELECT pvm as Date, open as Open, high as High, low as Low, close as Close, volume as Volume 
                FROM osakedata 
                WHERE osake = ? 
                ORDER BY pvm
            """,
                osake_conn,
                params=[ticker],
            )

            if df.empty:
                logging.warning(f"Ei hintadataa osakkeelle {ticker}")
                return

        except Exception as e:
            logging.warning(f"Virhe haettaessa dataa osakkeelle {ticker}: {e}")
            return

        # Konvertoi päivämäärät ja aseta indeksiksi
        df["Date"] = pd.to_datetime(df["Date"])
        df.set_index("Date", inplace=True)

        # Hae indeksitiedot kerralla
        spx_df = self._get_index_data(
            osake_conn, "^GSPC", df.index.min(), df.index.max()
        )
        ndx_df = self._get_index_data(
            osake_conn, "^NDX", df.index.min(), df.index.max()
        )

        # Laske liukuvat keskiarvot kerralla
        df["ma5"] = df["Close"].rolling(window=5).mean()
        df["ma10"] = df["Close"].rolling(window=10).mean()
        df["ma20"] = df["Close"].rolling(window=20).mean()
        df["ma50"] = df["Close"].rolling(window=50).mean()
        df["ma200"] = df["Close"].rolling(window=200).mean()

        # Laske volyymien 100pv keskiarvo
        df["volume_100d_avg"] = df["Volume"].rolling(window=100).mean()

        # Prosessoi löydökset batcheissa
        batch_size = 100
        for i in range(0, len(findings), batch_size):
            batch = findings[i : i + batch_size]
            self._process_findings_batch(
                ticker, batch, df, spx_df, ndx_df, results_conn
            )

    def _get_index_data(
        self, osake_conn, symbol: str, start_date, end_date
    ) -> pd.DataFrame:
        """Hae indeksidata annetulta aikaväliltä"""
        try:
            df = pd.read_sql(
                """
                SELECT pvm as Date, close as Close 
                FROM osakedata 
                WHERE osake = ? AND pvm BETWEEN ? AND ?
                ORDER BY pvm
            """,
                osake_conn,
                params=[
                    symbol,
                    start_date.strftime("%Y-%m-%d"),
                    end_date.strftime("%Y-%m-%d"),
                ],
            )

            if not df.empty:
                df["Date"] = pd.to_datetime(df["Date"])
                df.set_index("Date", inplace=True)

            return df

        except Exception as e:
            logging.warning(f"Virhe haettaessa indeksidataa {symbol}: {e}")
            return pd.DataFrame()

    def _process_findings_batch(
        self,
        ticker: str,
        findings_batch: List[Tuple[str, str]],
        df: pd.DataFrame,
        spx_df: pd.DataFrame,
        ndx_df: pd.DataFrame,
        results_conn,
    ):
        """Prosessoi löydösten batch optimoidusti"""

        batch_data = []

        for date_str, candle in findings_batch:
            try:
                date = pd.to_datetime(date_str)

                if date not in df.index:
                    continue

                # Laske kaikki arvot tälle löydökselle
                row_data = self._calculate_all_values(
                    ticker, date, candle, df, spx_df, ndx_df
                )

                if row_data:
                    batch_data.append(row_data)

            except Exception as e:
                logging.warning(f"Virhe prosessoitaessa {ticker} {date_str}: {e}")
                continue

        # Tallenna batch kerralla
        if batch_data:
            results_conn.executemany(
                """
                INSERT OR REPLACE INTO excel_staging 
                (osake, pvm, kynttila, 
                 t_1_alin, t_1_ylin, t_1_bodi, t_1_bodi_colour,
                 t0_alin, t0_ylin, t0_bodi, t0_bodi_colour, 
                 t1_alin, t1_ylin, t1_bodi, t1_bodi_colour,
                 t_5, t_10, t_20, 
                 t5, t10, t20, t40, t60, t252,
                 vol_5, vol_10, vol_20, vol_60, vol_252,
                 vol_avg_10, vol_avg_20, vol_avg_60,
                 vol_t_1_vs_avg10, vol_t0_vs_avg10, vol_t1_vs_avg10,
                 vol_t_1_vs_avg20, vol_t0_vs_avg20, vol_t1_vs_avg20, vol_spike,
                 ma_5, ma_10, ma_20, ma_50, ma_100, ma_200,
                 dist_ma_5, dist_ma_10, dist_ma_20, dist_ma_50, dist_ma_100, dist_ma_200,
                 ma_5_slope, ma_10_slope, ma_20_slope, ma_50_slope, ma_200_slope,
                 sp500_t_1, sp500_t0, sp500_t1, sp500_ma_20, sp500_ma_50, sp500_ma_200,
                 sp500_change_t0, sp500_change_t1, sp500_vs_ma20, sp500_vs_ma50, sp500_vs_ma200,
                 nasdaq_t_1, nasdaq_t0, nasdaq_t1, nasdaq_ma_20, nasdaq_ma_50, nasdaq_ma_200,
                 nasdaq_change_t0, nasdaq_change_t1, nasdaq_vs_ma20, nasdaq_vs_ma50, nasdaq_vs_ma200)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 
                       ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 
                       ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
                batch_data,
            )

    def _calculate_all_values(
        self,
        ticker: str,
        date: pd.Timestamp,
        candle: str,
        df: pd.DataFrame,
        spx_df: pd.DataFrame,
        ndx_df: pd.DataFrame,
    ) -> Optional[Tuple]:
        """Laske kaikki Excel-sarakkeiden arvot SARAKKEET_DOKUMENTAATIO.md mukaan (78 saraketta)"""

        try:
            # Etsi rivi DataFramesta
            if date not in df.index:
                return None

            # Hakemuksen helpottamiseksi indeksiarvo
            date_idx = df.index.get_loc(date)
            total_rows = len(df)

            # PERUSTIEDOT (1-3)
            # 1. osake = ticker
            # 2. pvm = date (string muodossa)
            # 3. kynttila = candle (integer)

            # T0 perusdata (käännekynttilä)
            row = df.iloc[date_idx]
            t0_open = row["Open"]
            t0_high = row["High"]
            t0_low = row["Low"]
            t0_close = row["Close"]
            t0_volume = row["Volume"]

            if t0_low is None or t0_low <= 0:
                return None

            # KYNTTILÄDETALJIT (4-15)
            # T-1 Edellinen päivä (4-7)
            def get_candle_values(offset):
                idx = date_idx + offset
                if idx < 0 or idx >= total_rows:
                    return None, None, None, None
                r = df.iloc[idx]
                alin_norm = (r["Low"] / t0_low) * 100.0  # Normalisoitu t0_alin:lla
                ylin_norm = (r["High"] / t0_low) * 100.0  # Normalisoitu t0_alin:lla
                bodi_pct = (
                    abs(r["Close"] - r["Open"]) / r["Open"] * 100.0
                )  # % avauksesta
                bodi_colour = 1 if r["Close"] > r["Open"] else 0  # 1=vihreä, 0=punainen
                return alin_norm, ylin_norm, bodi_pct, bodi_colour

            # 4-7: T-1 tiedot
            t_1_alin, t_1_ylin, t_1_bodi, t_1_bodi_colour = get_candle_values(-1)

            # 8-11: T0 tiedot (käännekynttilä)
            t0_alin_norm = 100.0  # Aina 100.00 (normalisointipohja)
            t0_ylin_norm = (t0_high / t0_low) * 100.0
            t0_bodi = abs(t0_close - t0_open) / t0_open * 100.0
            t0_bodi_colour = 1 if t0_close > t0_open else 0

            # 12-15: T1 tiedot
            t1_alin, t1_ylin, t1_bodi, t1_bodi_colour = get_candle_values(1)

            # HINNAT (16-24)
            # Historialliset hinnat (16-18) - normalisoitu t0_alin:lla
            def get_normalized_price(offset):
                idx = date_idx + offset
                if idx < 0 or idx >= total_rows:
                    return None
                close_val = df.iloc[idx]["Close"]
                return (close_val / t0_low) * 100.0 if close_val is not None else None

            t_5 = get_normalized_price(-5)  # 16. t_5
            t_10 = get_normalized_price(-10)  # 17. t_10
            t_20 = get_normalized_price(-20)  # 18. t_20

            # Tulevaisuuden hinnat (19-24) - normalisoitu t0_alin:lla
            t5 = get_normalized_price(5)  # 19. t5
            t10 = get_normalized_price(10)  # 20. t10
            t20 = get_normalized_price(20)  # 21. t20
            t40 = get_normalized_price(40)  # 22. t40
            t60 = get_normalized_price(60)  # 23. t60
            t252 = get_normalized_price(252)  # 24. t252

            # VOLATILITEETTI (25-29) - annualisoitu
            def calc_volatility(days):
                start_idx = max(0, date_idx - days + 1)
                end_idx = date_idx + 1
                prices = df.iloc[start_idx:end_idx]["Close"].dropna()
                if len(prices) < 2:
                    return None
                returns = prices.pct_change().dropna()
                if len(returns) < 1:
                    return None
                std_dev = returns.std()
                # Annualisoi: std * sqrt(252)
                return std_dev * (252**0.5) * 100.0  # Prosentteina

            vol_5 = calc_volatility(5)  # 25. vol_5
            vol_10 = calc_volatility(10)  # 26. vol_10
            vol_20 = calc_volatility(20)  # 27. vol_20
            vol_60 = calc_volatility(60)  # 28. vol_60
            vol_252 = calc_volatility(252)  # 29. vol_252

            # VOLYYMIT (30-39)
            # Keskivolyymit (30-32)
            def calc_avg_volume(days):
                start_idx = max(0, date_idx - days + 1)
                end_idx = date_idx + 1
                volumes = df.iloc[start_idx:end_idx]["Volume"].dropna()
                return volumes.mean() if not volumes.empty else None

            vol_avg_10 = calc_avg_volume(10)  # 30. vol_avg_10
            vol_avg_20 = calc_avg_volume(20)  # 31. vol_avg_20
            vol_avg_60 = calc_avg_volume(60)  # 32. vol_avg_60

            # Volyymisuhteet (33-38)
            def get_volume_ratio(offset, avg_days):
                idx = date_idx + offset
                if idx < 0 or idx >= total_rows:
                    return None
                vol = df.iloc[idx]["Volume"]
                if avg_days == 10:
                    avg_vol = vol_avg_10
                elif avg_days == 20:
                    avg_vol = vol_avg_20
                else:
                    return None
                return vol / avg_vol if avg_vol and avg_vol > 0 else None

            vol_t_1_vs_avg10 = get_volume_ratio(-1, 10)  # 33. vol_t_1_vs_avg10
            vol_t0_vs_avg10 = get_volume_ratio(0, 10)  # 34. vol_t0_vs_avg10
            vol_t1_vs_avg10 = get_volume_ratio(1, 10)  # 35. vol_t1_vs_avg10
            vol_t_1_vs_avg20 = get_volume_ratio(-1, 20)  # 36. vol_t_1_vs_avg20
            vol_t0_vs_avg20 = get_volume_ratio(0, 20)  # 37. vol_t0_vs_avg20
            vol_t1_vs_avg20 = get_volume_ratio(1, 20)  # 38. vol_t1_vs_avg20

            # 39. vol_spike (1 jos t0_vol > 2 * avg20, muuten 0)
            vol_spike = 1 if vol_t0_vs_avg20 and vol_t0_vs_avg20 > 2.0 else 0

            # LIUKUVAT KESKIARVOT (40-56)
            # MA:t (40-45) - normalisoitu t0_alin:lla
            def get_ma_normalized(days):
                if f"ma{days}" in df.columns:
                    ma_val = df.iloc[date_idx][f"ma{days}"]
                    return (ma_val / t0_low) * 100.0 if ma_val is not None else None
                return None

            ma_5 = get_ma_normalized(5)  # 40. ma_5
            ma_10 = get_ma_normalized(10)  # 41. ma_10
            ma_20 = get_ma_normalized(20)  # 42. ma_20
            ma_50 = get_ma_normalized(50)  # 43. ma_50
            ma_100 = get_ma_normalized(100)  # 44. ma_100
            ma_200 = get_ma_normalized(200)  # 45. ma_200

            # MA-etäisyydet (46-51) - prosentteina
            def calc_ma_distance(days):
                if f"ma{days}" in df.columns:
                    ma_val = df.iloc[date_idx][f"ma{days}"]
                    if ma_val and ma_val > 0:
                        return ((t0_close - ma_val) / ma_val) * 100.0
                return None

            dist_ma_5 = calc_ma_distance(5)  # 46. dist_ma_5
            dist_ma_10 = calc_ma_distance(10)  # 47. dist_ma_10
            dist_ma_20 = calc_ma_distance(20)  # 48. dist_ma_20
            dist_ma_50 = calc_ma_distance(50)  # 49. dist_ma_50
            dist_ma_100 = calc_ma_distance(100)  # 50. dist_ma_100
            dist_ma_200 = calc_ma_distance(200)  # 51. dist_ma_200

            # MA-kulmakertoimet (52-56)
            def calc_ma_slope(days):
                if f"ma{days}" in df.columns and date_idx >= days:
                    ma_today = df.iloc[date_idx][f"ma{days}"]
                    ma_past = df.iloc[date_idx - days][f"ma{days}"]
                    if ma_today is not None and ma_past is not None:
                        return (ma_today - ma_past) / days
                return None

            ma_5_slope = calc_ma_slope(5)  # 52. ma_5_slope
            ma_10_slope = calc_ma_slope(10)  # 53. ma_10_slope
            ma_20_slope = calc_ma_slope(20)  # 54. ma_20_slope
            ma_50_slope = calc_ma_slope(50)  # 55. ma_50_slope
            ma_200_slope = calc_ma_slope(200)  # 56. ma_200_slope

            # S&P 500 TIEDOT (57-67)
            def get_spx_value(offset):
                target_date = date + pd.Timedelta(days=offset)
                if target_date in spx_df.index:
                    return spx_df.loc[target_date, "Close"]
                return None

            sp500_t_1 = get_spx_value(-1)  # 57. sp500_t_1
            sp500_t0 = get_spx_value(0)  # 58. sp500_t0
            sp500_t1 = get_spx_value(1)  # 59. sp500_t1

            # S&P 500 MA:t (60-62)
            def calc_spx_ma(days):
                if len(spx_df) < days:
                    return None
                try:
                    # Etsi nykyinen päivä ja laske MA taaksepäin
                    if date in spx_df.index:
                        date_loc = spx_df.index.get_loc(date)
                        start_idx = max(0, date_loc - days + 1)
                        end_idx = date_loc + 1
                        ma_values = spx_df.iloc[start_idx:end_idx]["Close"].dropna()
                        if len(ma_values) >= days * 0.8:  # Vähintään 80% datasta
                            return ma_values.mean()
                except:
                    pass
                return None

            sp500_ma_20 = calc_spx_ma(20)  # 60. sp500_ma_20
            sp500_ma_50 = calc_spx_ma(50)  # 61. sp500_ma_50
            sp500_ma_200 = calc_spx_ma(200)  # 62. sp500_ma_200

            # S&P 500 muutokset (63-64)
            sp500_change_t0 = None  # 63. sp500_change_t0
            sp500_change_t1 = None  # 64. sp500_change_t1
            if sp500_t_1 and sp500_t0:
                sp500_change_t0 = ((sp500_t0 - sp500_t_1) / sp500_t_1) * 100.0
            if sp500_t0 and sp500_t1:
                sp500_change_t1 = ((sp500_t1 - sp500_t0) / sp500_t0) * 100.0

            # S&P 500 vs MA:t (65-67)
            sp500_vs_ma20 = None  # 65. sp500_vs_ma20
            sp500_vs_ma50 = None  # 66. sp500_vs_ma50
            sp500_vs_ma200 = None  # 67. sp500_vs_ma200

            if sp500_t0 and sp500_ma_20 and sp500_ma_20 > 0:
                sp500_vs_ma20 = ((sp500_t0 - sp500_ma_20) / sp500_ma_20) * 100.0
            if sp500_t0 and sp500_ma_50 and sp500_ma_50 > 0:
                sp500_vs_ma50 = ((sp500_t0 - sp500_ma_50) / sp500_ma_50) * 100.0
            if sp500_t0 and sp500_ma_200 and sp500_ma_200 > 0:
                sp500_vs_ma200 = ((sp500_t0 - sp500_ma_200) / sp500_ma_200) * 100.0

            # NASDAQ 100 TIEDOT (68-78)
            def get_ndx_value(offset):
                target_date = date + pd.Timedelta(days=offset)
                if target_date in ndx_df.index:
                    return ndx_df.loc[target_date, "Close"]
                return None

            nasdaq_t_1 = get_ndx_value(-1)  # 68. nasdaq_t_1
            nasdaq_t0 = get_ndx_value(0)  # 69. nasdaq_t0
            nasdaq_t1 = get_ndx_value(1)  # 70. nasdaq_t1

            # Nasdaq 100 MA:t (71-73)
            def calc_ndx_ma(days):
                if len(ndx_df) < days:
                    return None
                try:
                    # Etsi nykyinen päivä ja laske MA taaksepäin
                    if date in ndx_df.index:
                        date_loc = ndx_df.index.get_loc(date)
                        start_idx = max(0, date_loc - days + 1)
                        end_idx = date_loc + 1
                        ma_values = ndx_df.iloc[start_idx:end_idx]["Close"].dropna()
                        if len(ma_values) >= days * 0.8:  # Vähintään 80% datasta
                            return ma_values.mean()
                except:
                    pass
                return None

            nasdaq_ma_20 = calc_ndx_ma(20)  # 71. nasdaq_ma_20
            nasdaq_ma_50 = calc_ndx_ma(50)  # 72. nasdaq_ma_50
            nasdaq_ma_200 = calc_ndx_ma(200)  # 73. nasdaq_ma_200

            # Nasdaq 100 muutokset (74-75)
            nasdaq_change_t0 = None  # 74. nasdaq_change_t0
            nasdaq_change_t1 = None  # 75. nasdaq_change_t1
            if nasdaq_t_1 and nasdaq_t0:
                nasdaq_change_t0 = ((nasdaq_t0 - nasdaq_t_1) / nasdaq_t_1) * 100.0
            if nasdaq_t0 and nasdaq_t1:
                nasdaq_change_t1 = ((nasdaq_t1 - nasdaq_t0) / nasdaq_t0) * 100.0

            # Nasdaq 100 vs MA:t (76-78)
            nasdaq_vs_ma20 = None  # 76. nasdaq_vs_ma20
            nasdaq_vs_ma50 = None  # 77. nasdaq_vs_ma50
            nasdaq_vs_ma200 = None  # 78. nasdaq_vs_ma200

            if nasdaq_t0 and nasdaq_ma_20 and nasdaq_ma_20 > 0:
                nasdaq_vs_ma20 = ((nasdaq_t0 - nasdaq_ma_20) / nasdaq_ma_20) * 100.0
            if nasdaq_t0 and nasdaq_ma_50 and nasdaq_ma_50 > 0:
                nasdaq_vs_ma50 = ((nasdaq_t0 - nasdaq_ma_50) / nasdaq_ma_50) * 100.0
            if nasdaq_t0 and nasdaq_ma_200 and nasdaq_ma_200 > 0:
                nasdaq_vs_ma200 = ((nasdaq_t0 - nasdaq_ma_200) / nasdaq_ma_200) * 100.0

            # Palauta kaikki 78 arvoa oikeassa järjestyksessä

            # Muunna kynttilä-string numeroksi SARAKKEET_DOKUMENTAATIO.md mukaan
            candle_mapping = {
                "Hammer": 1,
                "Bullish Engulfing": 2,
                "Piercing Pattern": 3,
                "Morning Star": 4,
                "Falling Star": 5,
                "Hanging Man": 6,
                "Bearish Engulfing": 7,
            }
            candle_num = candle_mapping.get(candle, 0)  # 0 jos tuntematon

            return (
                # Perustiedot (1-3)
                ticker,
                date.strftime("%Y-%m-%d"),
                candle_num,
                # T-1 (4-7)
                t_1_alin,
                t_1_ylin,
                t_1_bodi,
                t_1_bodi_colour,
                # T0 (8-11)
                t0_alin_norm,
                t0_ylin_norm,
                t0_bodi,
                t0_bodi_colour,
                # T1 (12-15)
                t1_alin,
                t1_ylin,
                t1_bodi,
                t1_bodi_colour,
                # Historialliset hinnat (16-18)
                t_5,
                t_10,
                t_20,
                # Tulevat hinnat (19-24)
                t5,
                t10,
                t20,
                t40,
                t60,
                t252,
                # Volatiliteetti (25-29)
                vol_5,
                vol_10,
                vol_20,
                vol_60,
                vol_252,
                # Volyymit (30-39)
                vol_avg_10,
                vol_avg_20,
                vol_avg_60,
                vol_t_1_vs_avg10,
                vol_t0_vs_avg10,
                vol_t1_vs_avg10,
                vol_t_1_vs_avg20,
                vol_t0_vs_avg20,
                vol_t1_vs_avg20,
                vol_spike,
                # Liukuvat keskiarvot (40-56)
                ma_5,
                ma_10,
                ma_20,
                ma_50,
                ma_100,
                ma_200,
                dist_ma_5,
                dist_ma_10,
                dist_ma_20,
                dist_ma_50,
                dist_ma_100,
                dist_ma_200,
                ma_5_slope,
                ma_10_slope,
                ma_20_slope,
                ma_50_slope,
                ma_200_slope,
                # S&P 500 (57-67)
                sp500_t_1,
                sp500_t0,
                sp500_t1,
                sp500_ma_20,
                sp500_ma_50,
                sp500_ma_200,
                sp500_change_t0,
                sp500_change_t1,
                sp500_vs_ma20,
                sp500_vs_ma50,
                sp500_vs_ma200,
                # Nasdaq 100 (68-78)
                nasdaq_t_1,
                nasdaq_t0,
                nasdaq_t1,
                nasdaq_ma_20,
                nasdaq_ma_50,
                nasdaq_ma_200,
                nasdaq_change_t0,
                nasdaq_change_t1,
                nasdaq_vs_ma20,
                nasdaq_vs_ma50,
                nasdaq_vs_ma200,
            )

        except Exception as e:
            logging.error(f"Virhe laskennassa {ticker} {date}: {e}")
            return None

    def export_to_excel_fast(
        self,
        excel_path: str = "data/results.xlsx",
        limit_rows: int = None,
        ticker_filter: str = None,
        progress_callback=None,
    ) -> bool:
        """Nopea Excel-export staging-taulusta (78 saraketta dokumentaation mukaan)"""
        try:

            def update_progress(step: str, current: int, total: int):
                if progress_callback:
                    progress_callback(step, current, total)

            update_progress("Tarkistetaan cache", 10, 100)

            # Tarkista onko cache tuore, jos ei niin rakenna uudelleen
            if not self.is_cache_fresh():
                logging.info("Cache ei ole tuore, rakennetaan uudelleen...")
                update_progress("Rakennetaan cache", 20, 100)
                self.rebuild_staging_optimized(
                    limit_rows=limit_rows, ticker_filter=ticker_filter
                )

            update_progress("Luetaan data tietokannasta", 60, 100)

            # Lue kaikki data kerralla staging-taulusta
            with sqlite3.connect(self.results_db) as conn:
                # Rakenna SQL-kysely SARAKKEET_DOKUMENTAATIO.md mukaan (78 saraketta)
                base_query = """
                    SELECT 
                        -- Perustiedot (1-3)
                        osake, pvm, kynttila,
                        
                        -- T-1 Edellinen päivä (4-7)
                        t_1_alin, t_1_ylin, t_1_bodi, t_1_bodi_colour,
                        
                        -- T0 Käännekynttilä (8-11)
                        t0_alin, t0_ylin, t0_bodi, t0_bodi_colour,
                        
                        -- T1 Seuraava päivä (12-15)
                        t1_alin, t1_ylin, t1_bodi, t1_bodi_colour,
                        
                        -- Historialliset hinnat (16-18)
                        t_5, t_10, t_20,
                        
                        -- Tulevat hinnat (19-24)
                        t5, t10, t20, t40, t60, t252,
                        
                        -- Volatiliteetti (25-29)
                        vol_5, vol_10, vol_20, vol_60, vol_252,
                        
                        -- Volyymit (30-39)
                        vol_avg_10, vol_avg_20, vol_avg_60,
                        vol_t_1_vs_avg10, vol_t0_vs_avg10, vol_t1_vs_avg10,
                        vol_t_1_vs_avg20, vol_t0_vs_avg20, vol_t1_vs_avg20,
                        vol_spike,
                        
                        -- Liukuvat keskiarvot (40-56)
                        ma_5, ma_10, ma_20, ma_50, ma_100, ma_200,
                        dist_ma_5, dist_ma_10, dist_ma_20, dist_ma_50, dist_ma_100, dist_ma_200,
                        ma_5_slope, ma_10_slope, ma_20_slope, ma_50_slope, ma_200_slope,
                        
                        -- S&P 500 tiedot (57-67)
                        sp500_t_1, sp500_t0, sp500_t1,
                        sp500_ma_20, sp500_ma_50, sp500_ma_200,
                        sp500_change_t0, sp500_change_t1,
                        sp500_vs_ma20, sp500_vs_ma50, sp500_vs_ma200,
                        
                        -- Nasdaq 100 tiedot (68-78)
                        nasdaq_t_1, nasdaq_t0, nasdaq_t1,
                        nasdaq_ma_20, nasdaq_ma_50, nasdaq_ma_200,
                        nasdaq_change_t0, nasdaq_change_t1,
                        nasdaq_vs_ma20, nasdaq_vs_ma50, nasdaq_vs_ma200
                        
                    FROM excel_staging"""

                # Lisää WHERE-lauseke ticker-filtteriä varten
                conditions = []
                params = []

                if ticker_filter:
                    conditions.append("osake = ?")
                    params.append(ticker_filter)

                if conditions:
                    base_query += " WHERE " + " AND ".join(conditions)

                base_query += " ORDER BY osake, pvm"

                if limit_rows:
                    query = base_query + f" LIMIT {limit_rows}"
                else:
                    query = base_query

                if params:
                    df = pd.read_sql(query, conn, params=params)
                else:
                    df = pd.read_sql(query, conn)

            if df.empty:
                logging.warning("Ei dataa staging-taulussa")
                return False

            update_progress("Tallennetaan Excel-tiedosto", 80, 100)

            # Tallenna Excel-tiedostoon ILMAN suomalaista muotoilua (pidä numerot numeroina)
            Path(excel_path).parent.mkdir(parents=True, exist_ok=True)
            df.to_excel(excel_path, index=False, engine="openpyxl")

            # Muotoile numerosarakkeet (sarakkeet 3-78) kahdella desimaalilla
            update_progress("Muotoillaan numerosarakkeet", 95, 100)

            import openpyxl

            # Avaa Excel-tiedosto muotoilua varten
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active

            # Muotoile sarakkeet 3-78 (C-??), ohita sarakkeet 1-2 (A-B)
            for col in range(3, min(79, ws.max_column + 1)):  # Sarakkeet C-??
                col_letter = openpyxl.utils.get_column_letter(col)
                for row in range(2, ws.max_row + 1):  # Ohita otsikkorivi
                    cell = ws[f"{col_letter}{row}"]
                    if cell.value is not None and isinstance(cell.value, (int, float)):
                        # Käytä suomalaista muotoilua: pilkku desimaalimerkkinä, ei tuhaterotinta
                        cell.number_format = "0,00"

            # Tallenna muotoillut muutokset
            wb.save(excel_path)
            wb.close()

            update_progress("Valmis!", 100, 100)

            logging.info(f"Excel-tiedosto tallennettu: {excel_path} ({len(df)} riviä)")
            return True

        except Exception as e:
            logging.error(f"Virhe Excel-exportissa: {e}")
            return False

    def _format_finnish_number(self, value):
        """Muotoile numero suomalaiseen muotoon"""
        if value is None or pd.isna(value):
            return ""
        try:
            if isinstance(value, (int, float)):
                formatted = f"{float(value):.4f}".rstrip("0").rstrip(".")
                return formatted.replace(".", ",")
            return str(value).replace(".", ",")
        except:
            return str(value)

    def get_staging_stats(self) -> dict:
        """Hae staging-taulun tilastot"""
        try:
            with sqlite3.connect(self.results_db) as conn:
                stats = conn.execute(
                    """
                    SELECT 
                        COUNT(*) as total_rows,
                        COUNT(DISTINCT ticker) as unique_tickers,
                        MIN(date) as earliest_date,
                        MAX(date) as latest_date
                    FROM excel_staging
                """
                ).fetchone()

                last_update = conn.execute(
                    """
                    SELECT value FROM cache_metadata 
                    WHERE key = 'last_full_rebuild'
                """
                ).fetchone()

                return {
                    "total_rows": stats[0] if stats else 0,
                    "unique_tickers": stats[1] if stats else 0,
                    "earliest_date": stats[2] if stats else None,
                    "latest_date": stats[3] if stats else None,
                    "last_update": last_update[0] if last_update else None,
                }
        except Exception as e:
            logging.error(f"Virhe tilastojen haussa: {e}")
            return {}
