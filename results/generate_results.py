import sqlite3
import threading
import traceback
from pathlib import Path
from statistics import mean, pstdev
from typing import List, Optional, Tuple
import locale

import flet as ft
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Uusi optimoitu cache-järjestelmä
from .excel_cache import ExcelResultsCache


def _identify_columns(cur, table_name: str):
    cols = [r[1] for r in cur.execute(f"PRAGMA table_info({table_name})").fetchall()]
    lower = {c.lower(): c for c in cols}
    return cols, lower


def _calculate_relative_stdev(values: List[float]) -> Optional[float]:
    """Laskee suhteellisen standardipoikkeaman (% keskiarvosta)"""
    try:
        if len(values) < 2:
            return None
        values = [v for v in values if v is not None]
        if len(values) < 2:
            return None
        avg = mean(values)
        if avg == 0:
            return None
        std = pstdev(values)
        return (std / avg) * 100.0
    except Exception:
        return None


def _format_finnish_number(value):
    """Muotoilee numeron suomalaiseen muotoon: pilkku desimaalimerkkinä, ei tuhateroittimia"""
    if value is None or value == "":
        return ""
    try:
        if isinstance(value, (int, float)):
            # Pyöristetään 4 desimaaliin ja vaihdetaan piste pilkkuun
            formatted = f"{float(value):.4f}".rstrip("0").rstrip(".")
            return formatted.replace(".", ",")
        return str(value).replace(".", ",")
    except (ValueError, TypeError):
        return str(value)


def _create_excel_file(
    header: List[str],
    output_rows: List[List],
    excel_path: Path,
    downtrend_filter: bool = True,
):
    """Luo Excel-tiedoston suomalaisilla asetuksilla ja kauniilla muotoilulla"""
    try:
        # Luodaan DataFrame (säilytetään alkuperäiset numerot)
        df = pd.DataFrame(output_rows, columns=header)

        # Luodaan workbook ja worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Kynttilätulokset"

        # Suomalainen numeromuotoilu Excelille
        finnish_number_format = "#,##0.0000"
        finnish_number_format = (
            finnish_number_format.replace(",", " ").replace(".", ",").replace(" ", ".")
        )

        # Lisätään data
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                # Muotoilu otsikoille
                if r_idx == 1:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(
                        start_color="366092", end_color="366092", fill_type="solid"
                    )
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                # Numeromuotoilu numerotyyppisille soluille (sarakkeet 4 alkaen)
                elif r_idx > 1 and c_idx >= 4:  # Data-rivit, numerosarakkeet
                    if isinstance(value, (int, float)) and value is not None:
                        cell.number_format = "#,##0.00;[RED]-#,##0.00"

                # Muotoilu trendisarakkeelle jos käytössä
                elif downtrend_filter and c_idx == 3:  # Kynttilä-sarake
                    if "downtrend" in str(value).lower():
                        cell.fill = PatternFill(
                            start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"
                        )

        # Automaattinen sarakkeiden leveyden säätö
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 25)  # Max 25 merkkiä
            ws.column_dimensions[column_letter].width = adjusted_width

        # Tallennetaan
        excel_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(excel_path)
        return True

    except Exception as e:
        print(f"Virhe Excel-tiedoston luonnissa: {e}")
        return False


def _calculate_moving_average(
    df: pd.DataFrame, ccol: str, idx: int, days: int
) -> Optional[float]:
    """Laskee liukuvan keskiarvon annetusta indeksistä taaksepäin"""
    try:
        if idx - days + 1 < 0:
            return None
        subset = df.iloc[idx - days + 1 : idx + 1]
        values = [
            float(row[ccol]) for _, row in subset.iterrows() if pd.notna(row[ccol])
        ]
        if len(values) != days:
            return None
        return mean(values)
    except Exception:
        return None


def _is_in_downtrend(
    df: pd.DataFrame,
    ccol: str,
    vcol: str,
    idx: int,
    min_decline_percent: float = 3.0,
    use_ma_filter: bool = True,
    use_volume_filter: bool = False,
) -> bool:
    """Tarkistaa onko kynttilä laskutrendissä annettujen kriteerien mukaan"""
    try:
        # Tarvitaan vähintään 10 päivää historiaa
        if idx < 10:
            return False

        def safe_get(row_idx, col):
            if row_idx < 0 or row_idx >= len(df):
                return None
            try:
                val = df.iloc[row_idx][col]
                return float(val) if pd.notna(val) else None
            except Exception:
                return None

        # 1. Peruskriteeri: Porrastava lasku t-10 > t-5 > t-2 > t0
        t0 = safe_get(idx, ccol)
        t_2 = safe_get(idx - 2, ccol)
        t_5 = safe_get(idx - 5, ccol)
        t_10 = safe_get(idx - 10, ccol)

        if not all([t0, t_2, t_5, t_10]):
            return False

        if not (t_10 > t_5 > t_2 > t0):
            return False

        # 2. Minimalasku: vähintään X% laskua 10 päivässä
        decline_percent = ((t_10 - t0) / t_10) * 100
        if decline_percent < min_decline_percent:
            return False

        # 3. Liukuva keskiarvo -suodatin (valinnainen)
        if use_ma_filter:
            ma5 = _calculate_moving_average(df, ccol, idx, 5)
            ma10 = _calculate_moving_average(df, ccol, idx, 10)

            if ma5 is None or ma10 is None:
                return False

            # Kurssi alle MA(10) ja MA(5) < MA(10)
            if not (t0 < ma10 and ma5 < ma10):
                return False

        # 4. Volyymi-suodatin (valinnainen)
        if use_volume_filter:
            try:
                # Keskivolyymi viimeisen 5 päivän ajalta
                recent_volumes = []
                for i in range(max(0, idx - 4), idx + 1):
                    vol = safe_get(i, vcol)
                    if vol and vol > 0:
                        recent_volumes.append(vol)

                # Keskivolyymi 20 päivän historiasta (päivät -25 ... -5)
                historical_volumes = []
                for i in range(max(0, idx - 25), max(0, idx - 4)):
                    vol = safe_get(i, vcol)
                    if vol and vol > 0:
                        historical_volumes.append(vol)

                if not recent_volumes or not historical_volumes:
                    return False

                recent_avg = mean(recent_volumes)
                historical_avg = mean(historical_volumes)

                # Volyymi vähintään 1.2x normaalia
                if recent_avg < 1.2 * historical_avg:
                    return False

            except Exception:
                # Jos volyymitarkistus epäonnistuu, hyväksytään kuitenkin
                pass

        return True

    except Exception:
        return False


def _calculate_volume_ratio(
    df: pd.DataFrame, vcol: str, idx: int, days_range: Tuple[int, int]
) -> Optional[float]:
    """Laskee volyymin suhteen 100 päivän keskiarvoon"""
    try:
        start_offset, end_offset = days_range
        start_idx = idx + start_offset
        end_idx = idx + end_offset

        if start_idx < 0 or end_idx >= len(df):
            return None

        # Laske keskivolyymi annetulta aikaväliltä
        subset = df.iloc[start_idx : end_idx + 1]
        volumes = [
            float(row[vcol])
            for _, row in subset.iterrows()
            if pd.notna(row[vcol]) and row[vcol] > 0
        ]
        if not volumes:
            return None
        avg_volume = mean(volumes)

        # Laske 100 päivän keskiarvo ennen indeksiä
        hundred_start = max(0, idx - 100)
        hundred_subset = df.iloc[hundred_start:idx]
        hundred_volumes = [
            float(row[vcol])
            for _, row in hundred_subset.iterrows()
            if pd.notna(row[vcol]) and row[vcol] > 0
        ]
        if not hundred_volumes:
            return None
        hundred_avg = mean(hundred_volumes)

        if hundred_avg == 0:
            return None

        return avg_volume / hundred_avg
    except Exception:
        return None


def _get_index_data(
    oconn, ticker: str, date: str, offset: int, data_type: str = "close"
) -> Optional[float]:
    """Hakee indeksi-tietoja tietokannasta"""
    try:
        # Normalisoi ticker: isot kirjaimet, trimmattu
        ticker = (ticker or "").strip().upper()

        # Hae indeksin data
        df = pd.read_sql_query(
            "SELECT pvm, open, high, low, close FROM osakedata WHERE osake = ? ORDER BY pvm ASC",
            oconn,
            params=(ticker,),
        )
        if df.empty:
            return None

        df["pvm"] = pd.to_datetime(df["pvm"]).dt.strftime("%Y-%m-%d")
        date_to_idx = {str(row["pvm"]): idx for idx, row in df.iterrows()}

        if date not in date_to_idx:
            return None

        target_idx = date_to_idx[date] + offset
        if target_idx < 0 or target_idx >= len(df):
            return None

        row = df.iloc[target_idx]
        return float(row[data_type]) if pd.notna(row[data_type]) else None
    except Exception:
        return None


def _build_output_rows(
    analysis_db: Path,
    osake_db: Path,
    downtrend_filter: bool = True,
    min_decline_percent: float = 3.0,
    use_ma_filter: bool = True,
    use_volume_filter: bool = False,
    progress_callback=None,
    ticker_filter: Optional[str | list] = None,
    pattern_filter: Optional[list] = None,
):
    """Synchronous builder for output rows according to spec.

    Returns (header, output_rows).

    Args:
        downtrend_filter: Jos True, suodatetaan vain laskutrendien kynttilät
        min_decline_percent: Minimalasku prosentteina
        use_ma_filter: Käytetäänkö liukuva keskiarvo -suodatinta
        use_volume_filter: Käytetäänkö volyymi-suodatinta
        progress_callback: Callback-funktio edistymisen raportointiin
        ticker_filter: Jos annettu, rajaa tulokset tiettyyn tickeriin tai ticker-listaan
        pattern_filter: Jos annettu, rajaa tulokset vain valittuihin kynttiläkuvioihin
    """

    # Candlestick pattern to integer mapping
    # Kynttilöiden numerointi:
    # 1 = Hammer
    # 2 = Bullish Engulfing
    # 3 = Piercing Pattern
    # 4 = Three White Soldiers
    # 5 = Morning Star
    # 6 = Dragonfly Doji
    # 7 = Bullish Divergence
    # 8 = Bearish Divergence
    # 0 = downtrend (ei kynttilämalli)
    CANDLE_MAPPING = {
        "Hammer": 1,
        "Bullish Engulfing": 2,
        "Piercing Pattern": 3,
        "Three White Soldiers": 4,
        "Morning Star": 5,
        "Dragonfly Doji": 6,
        "Bullish Divergence": 7,
        "Bearish Divergence": 8,
        "downtrend": 0,  # Ei kynttilämalli, vain laskutrendi-indikaattori
    }

    # --- read analysis rows ---
    with sqlite3.connect(analysis_db) as aconn:
        acur = aconn.cursor()
        tbls = [
            r[0]
            for r in acur.execute(
                "SELECT name FROM sqlite_master WHERE type='table'"
            ).fetchall()
        ]
        candidates = ["analysis_findings", "analysis", "findings", "analysis_rows"]
        table_name = next((c for c in candidates if c in tbls), None)
        if table_name is None:
            # fallback heuristic
            for t in tbls:
                info = acur.execute(f"PRAGMA table_info({t})").fetchall()
                cols = [r[1].lower() for r in info]
                if any(x in cols for x in ("date", "pvm")) and any(
                    x in cols for x in ("ticker", "osake", "symbol")
                ):
                    table_name = t
                    break
        if table_name is None:
            raise RuntimeError("analysis table not found in analysis.db")

        info = acur.execute(f"PRAGMA table_info({table_name})").fetchall()
        col_names = [r[1] for r in info]
        lower = {c.lower(): c for c in col_names}
        date_col = lower.get("date") or lower.get("pvm")
        ticker_col = lower.get("ticker") or lower.get("osake") or lower.get("symbol")
        candle_col = (
            lower.get("candle")
            or lower.get("kynttila")
            or lower.get("pattern")
            or lower.get("pattern_name")
        )
        signal_strength_col = lower.get("signal_strength") or lower.get("strength")
        rsi14_col = lower.get("rsi14")

        if not date_col or not ticker_col:
            raise RuntimeError(
                f"Cannot find date/ticker columns in {table_name}: {col_names}"
            )

        q = f'SELECT "{date_col}", "{ticker_col}"'
        if candle_col:
            q += f', "{candle_col}"'
        if signal_strength_col:
            q += f', "{signal_strength_col}"'
        if rsi14_col:
            q += f', "{rsi14_col}"'
        q += f' FROM "{table_name}"'
        acur.execute(q)
        rows = acur.fetchall()

    # Define the complete header according to final specification (81 columns)
    # HUOM: Näitä otsikoita ei saa muuttaa, jotta Excelin sarakejärjestys säilyy.
    header = [
        # Perustieto (1-4)
        "osake",  # 1. osake
        "date",  # 2. date
        "kynttila",  # 3. kynttila
        "vahvuus",  # 4. vahvuus (signal_strength)
        # Kynttilätiedot (5-16)
        "t_1_alin",  # 5. t_1_alin
        "t_1_ylin",  # 6. t_1_ylin
        "t_1_bodi",  # 7. t_1_bodi
        "t_1_bodi_colour",  # 8. t_1_bodi_colour
        "t0_alin",  # 9. t0_alin (aina 100)
        "t0_ylin",  # 10. t0_ylin
        "t0_bodi",  # 11. t0_bodi
        "t0_bodi_colour",  # 12. t0_bodi_colour
        "t1_alin",  # 13. t1_alin
        "t1_ylin",  # 14. t1_ylin
        "t1_bodi",  # 15. t1_bodi
        "t1_bodi_colour",  # 16. t1_bodi_colour
        # Historialliset päätöskurssit (17-21)
        "t_2",  # 17. t_2
        "t_5",  # 18. t_5
        "t_10",  # 19. t_10
        "t_15",  # 20. t_15
        "t_20",  # 21. t_20
        # Volatiliteetti (22-26)
        "t_2_hajonta",  # 22. t_2_hajonta
        "t_5_hajonta",  # 23. t_5_hajonta
        "t_10_hajonta",  # 24. t_10_hajonta
        "t_15_hajonta",  # 25. t_15_hajonta
        "t_20_hajonta",  # 26. t_20_hajonta
        # Tulevat päätöskurssit (27-30)
        "t2",  # 27. t2
        "t5",  # 28. t5
        "t10",  # 29. t10
        "t20",  # 30. t20
        # Volyymit (31-40)
        "t_2_volyymi",  # 31. t_2_volyymi
        "t_5_volyymi",  # 32. t_5_volyymi
        "t_10_volyymi",  # 33. t_10_volyymi
        "t_15_volyymi",  # 34. t_15_volyymi
        "t_20_volyymi",  # 35. t_20_volyymi
        "t0_volyymi",  # 36. t0_volyymi
        "t2_volyymi",  # 37. t2_volyymi
        "t5_volyymi",  # 38. t5_volyymi
        "t10_volyymi",  # 39. t10_volyymi
        "t20_volyymi",  # 40. t20_volyymi
        # Liukuvat keskiarvot (41-55)
        "t_2_5p_liukuva",  # 41. t_2_5p_liukuva
        "t_2_10p_liukuva",  # 42. t_2_10p_liukuva
        "t_2_20p_liukuva",  # 43. t_2_20p_liukuva
        "t_5_5p_liukuva",  # 44. t_5_5p_liukuva
        "t_5_10p_liukuva",  # 45. t_5_10p_liukuva
        "t_5_20p_liukuva",  # 46. t_5_20p_liukuva
        "t_10_5p_liukuva",  # 47. t_10_5p_liukuva
        "t_10_10p_liukuva",  # 48. t_10_10p_liukuva
        "t_10_20p_liukuva",  # 49. t_10_20p_liukuva
        "t_15_5p_liukuva",  # 50. t_15_5p_liukuva
        "t_15_10p_liukuva",  # 51. t_15_10p_liukuva
        "t_15_20p_liukuva",  # 52. t_15_20p_liukuva
        "t_20_5p_liukuva",  # 53. t_20_5p_liukuva
        "t_20_10p_liukuva",  # 54. t_20_10p_liukuva
        "t_20_20p_liukuva",  # 55. t_20_20p_liukuva
        "t0_50p_liukuva",  # 56. t0_50p_liukuva
        "t0_200p_liukuva",  # 57. t0_200p_liukuva
        # S&P 500 index (58-68)
        "SPX_0",  # 58. SPX_0
        "SPX_2",  # 59. SPX_2
        "SPX_5",  # 60. SPX_5
        "SPX_10",  # 61. SPX_10
        "SPX_15",  # 62. SPX_15
        "SPX_20",  # 63. SPX_20
        "SPX2",  # 64. SPX2
        "SPX5",  # 65. SPX5
        "SPX10",  # 66. SPX10
        "SPX15",  # 67. SPX15
        "SPX20",  # 68. SPX20
        # Nasdaq 100 index (69-79)
        "NDX_0",  # 69. NDX_0
        "NDX_2",  # 70. NDX_2
        "NDX_5",  # 71. NDX_5
        "NDX_10",  # 72. NDX_10
        "NDX_15",  # 73. NDX_15
        "NDX_20",  # 74. NDX_20
        "NDX2",  # 75. NDX2
        "NDX5",  # 76. NDX5
        "NDX10",  # 77. NDX10
        "NDX15",  # 78. NDX15
        "NDX20",  # 79. NDX20
        "RSI14_t0",  # 80. RSI14_t0
        "t0_close_norm",  # 81. Normalisoitu t0_close
        "Bearish Divergence",  # 82. Bearish Divergence vahvuus (1.00-3.00, 2 desimaalia) jos löytyy t0/t-1/t-2/t-3, muuten 0
        "Bullish Divergence",  # 83. Bullish Divergence vahvuus (1.00-3.00, 2 desimaalia) jos löytyy t0/t-1/t-2/t-3, muuten 0
    ]

    if not rows:
        return header, []

    # Ticker-filtteri voi olla merkkijono tai lista
    ticker_filter_set = None
    if ticker_filter:
        if isinstance(ticker_filter, list):
            ticker_filter_set = {t.upper() for t in ticker_filter if t}
        else:
            ticker_filter_set = {ticker_filter.upper()}

    # Pattern-filtteri - muunna lista setiksi vertailua varten
    pattern_filter_set = None
    if pattern_filter:
        pattern_filter_set = {p for p in pattern_filter if p}

    by_ticker = {}
    for rec in rows:
        # Handle different numbers of columns based on what was selected
        if len(rec) >= 5:  # New format with signal_strength and rsi14
            date, ticker, candle, signal_strength, rsi14 = rec[:5]
        elif len(rec) >= 4:  # Format with signal_strength but no rsi14
            date, ticker, candle, signal_strength = rec[:4]
            rsi14 = None
        elif len(rec) == 3:  # Legacy format
            date, ticker, candle = rec
            signal_strength = None
            rsi14 = None
        elif len(rec) == 2:  # Legacy format without candle
            date, ticker = rec
            candle = ""
            signal_strength = None
            rsi14 = None
        else:
            continue

        if not ticker:
            continue

        ticker_str = str(ticker)
        if ticker_filter_set and ticker_str.upper() not in ticker_filter_set:
            continue

        # Suodata pattern:in mukaan jos pattern_filter on asetettu
        if pattern_filter_set and candle:
            if str(candle) not in pattern_filter_set:
                continue

        by_ticker.setdefault(ticker_str, []).append(
            (str(date), candle, signal_strength, rsi14)
        )

    output_rows = []

    # --- read osakedata per ticker ---
    with sqlite3.connect(osake_db) as oconn:
        ocur = oconn.cursor()
        cols, lower = _identify_columns(ocur, "osakedata")
        tcol = lower.get("osake") or lower.get("ticker") or lower.get("symbol")
        pcol = lower.get("pvm") or lower.get("date")
        ocol = lower.get("open")
        hcol = lower.get("high")
        lcol = lower.get("low")
        ccol = lower.get("close")
        vcol = lower.get("volume")

        if (
            not tcol
            or not pcol
            or not ocol
            or not hcol
            or not lcol
            or not ccol
            or not vcol
        ):
            raise RuntimeError("osakedata table missing expected column names")

        total_tickers = len(by_ticker)
        for ticker_idx, (ticker, items) in enumerate(by_ticker.items()):
            # Progress indicator
            progress = (
                0.2 + (ticker_idx / total_tickers) * 0.6
            )  # 20% - 80% of total progress

            if ticker_idx % 10 == 0:
                print(
                    f"📊 Käsitellään ticker {ticker_idx + 1}/{total_tickers}: {ticker}"
                )

            if progress_callback:
                progress_callback(
                    f"📊 Käsitellään osake {ticker_idx + 1}/{total_tickers}: {ticker}",
                    progress,
                )

            try:
                df = pd.read_sql_query(
                    f'SELECT * FROM osakedata WHERE "{tcol}" = ? ORDER BY "{pcol}" ASC',
                    oconn,
                    params=(ticker,),
                )
            except Exception:
                df = pd.read_sql_query(
                    f'SELECT * FROM osakedata ORDER BY "{pcol}" ASC', oconn
                )
                if tcol in df.columns:
                    df = df[df[tcol] == ticker]

            if df.empty:
                continue

            # Check if this ticker is an index (starts with ^)
            is_index = ticker.startswith("^")

            try:
                df[pcol] = pd.to_datetime(df[pcol]).dt.strftime("%Y-%m-%d")
            except Exception:
                pass

            df = df.reset_index(drop=True)
            date_to_idx = {str(r[pcol]): idx for idx, r in df.iterrows()}

            # RSI lasketaan nyt analysis-vaiheessa ja luetaan analysis.db:stä
            # Ei tarvita erillistä RSI-laskentaa täällä

            def safe_get(row, col):
                try:
                    v = row[col]
                    if pd.isna(v):
                        return None
                    return float(v)
                except Exception:
                    return None

            for date, candle, signal_strength, rsi14_from_db in items:
                if date not in date_to_idx:
                    continue
                idx = date_to_idx[date]

                # Check if we have enough data for all calculations (except t200_200p_liukuva)
                if idx < 20 or idx + 20 >= len(df):
                    continue

                # Get divergence flags for t0, t-1, t-2, t-3
                # Find the 4 trading days: t0 (current), t-1, t-2, t-3
                check_dates = []
                for offset in [0, -1, -2, -3]:
                    check_idx = idx + offset
                    if 0 <= check_idx < len(df):
                        check_date = str(df.iloc[check_idx][pcol])
                        check_dates.append(check_date)

                # Hae divergenssit analysis.db:stä (t0, t-1, t-2, t-3)
                from analysis.database_manager import DatabaseManager

                db_mgr = DatabaseManager(db_path=str(analysis_db))
                bearish_div, bullish_div = db_mgr.get_divergences_for_dates(
                    ticker=ticker, dates=check_dates
                )
                db_mgr.close()

                # Laskutrendi-suodatus
                if downtrend_filter:
                    if not _is_in_downtrend(
                        df,
                        ccol,
                        vcol,
                        idx,
                        min_decline_percent,
                        use_ma_filter,
                        use_volume_filter,
                    ):
                        continue  # Ohita kynttilät jotka eivät ole laskutrendissä

                # Get t0 values for normalization
                r0 = df.loc[idx]
                t0_low = safe_get(r0, lcol)
                t0_high = safe_get(r0, hcol)
                t0_open = safe_get(r0, ocol)
                t0_close = safe_get(r0, ccol)

                if t0_low is None or t0_low <= 0 or t0_close is None or t0_close <= 0:
                    continue

                t0_close_norm = (t0_close / t0_low * 100) if t0_low else None

                # Helper function to calculate body percentage and color
                def calc_candle_details(row_data):
                    if row_data is None:
                        return None, None, None, None

                    low = safe_get(row_data, lcol)
                    high = safe_get(row_data, hcol)
                    open_val = safe_get(row_data, ocol)
                    close_val = safe_get(row_data, ccol)

                    if any(x is None for x in [low, high, open_val, close_val]):
                        return None, None, None, None

                    # Normalize to t0_low
                    norm_low = (low / t0_low * 100) if t0_low > 0 else None
                    norm_high = (high / t0_low * 100) if t0_low > 0 else None

                    # Body percentage of total candle
                    candle_range = high - low
                    body_size = abs(close_val - open_val)
                    body_percent = (
                        (body_size / candle_range * 100) if candle_range > 0 else 0
                    )

                    # Color: 1=green (close > open), 0=red (close <= open)
                    color = 1 if close_val > open_val else 0

                    return norm_low, norm_high, body_percent, color

                # Calculate detailed candle data for t-1, t0, t1
                r_m1 = df.loc[idx - 1] if idx > 0 else None
                r1 = df.loc[idx + 1] if idx + 1 < len(df) else None

                # T-1 (previous day)
                t_1_alin, t_1_ylin, t_1_bodi, t_1_bodi_colour = calc_candle_details(
                    r_m1
                )

                # T0 (current day - candle day)
                t0_alin, t0_ylin, t0_bodi, t0_bodi_colour = calc_candle_details(r0)

                # T1 (next day)
                t1_alin, t1_ylin, t1_bodi, t1_bodi_colour = calc_candle_details(r1)

                # Calculate new historical and future prices (normalized)
                def get_normalized_close(offset):
                    target_idx = idx + offset
                    if target_idx < 0 or target_idx >= len(df):
                        return None
                    target_row = df.loc[target_idx]
                    close_val = safe_get(target_row, ccol)

                    if close_val is None:
                        return None

                    if is_index:
                        # Indices: normalize to t0_close=100
                        return (
                            (close_val / t0_close * 100)
                            if t0_close and t0_close > 0
                            else None
                        )
                    else:
                        # Stocks: normalize to t0_alin=100 (t0_low)
                        return (
                            (close_val / t0_low * 100)
                            if t0_low and t0_low > 0
                            else None
                        )

                t_2 = get_normalized_close(-2)
                t_5 = get_normalized_close(-5)
                t_10 = get_normalized_close(-10)
                t_15 = get_normalized_close(-15)
                t_20 = get_normalized_close(-20)

                t2 = get_normalized_close(2)
                t5 = get_normalized_close(5)
                t10 = get_normalized_close(10)
                t20 = get_normalized_close(20)

                # Calculate volatility (standard deviation, NO NORMALIZATION)
                def calc_volatility(days_back):
                    # Muutettu: lasketaan hajonta t0:n edeltäviltä päiviltä (ei t0 mukana), normalisoiduilla arvoilla
                    if idx - days_back < 0:
                        return None
                    start_idx = idx - days_back
                    end_idx = idx - 1  # Ei sisällä t0
                    subset = df.iloc[start_idx : end_idx + 1]
                    t0_low = safe_get(r0, lcol)
                    if t0_low is None or t0_low == 0:
                        return None
                    values = [safe_get(row, ccol) for _, row in subset.iterrows()]
                    values = [v for v in values if v is not None]
                    if len(values) < 2:
                        return None
                    # Normalisoi arvot t0_low:lla ja kerro 100
                    norm_values = [(v / t0_low) * 100 for v in values]
                    try:
                        return pstdev(norm_values)
                    except Exception:
                        return None

                t_2_hajonta = calc_volatility(2)
                t_5_hajonta = calc_volatility(5)
                t_10_hajonta = calc_volatility(10)
                t_15_hajonta = calc_volatility(15)
                t_20_hajonta = calc_volatility(20)

                # Calculate volume ratios according to specification
                def calc_volume_ratio_spec(days_count, offset_start):
                    """
                    Calculate volume ratio as specified:
                    - days_count: number of days to average
                    - offset_start: starting offset from t0 (negative for past, positive for future)
                    """
                    try:
                        # Calculate volume average for the specified period
                        start_idx = idx + offset_start
                        end_idx = start_idx + days_count - 1

                        if start_idx < 0 or end_idx >= len(df):
                            return None

                        subset = df.iloc[start_idx : end_idx + 1]
                        volumes = [safe_get(row, vcol) for _, row in subset.iterrows()]
                        volumes = [v for v in volumes if v is not None and v > 0]

                        if not volumes:
                            return None

                        period_avg = mean(volumes)

                        # Calculate 100-day average ending at t-1
                        hundred_start = max(0, idx - 100)
                        hundred_end = idx - 1  # End at t-1

                        if hundred_end < hundred_start:
                            return None

                        hundred_subset = df.iloc[hundred_start : hundred_end + 1]
                        hundred_volumes = [
                            safe_get(row, vcol) for _, row in hundred_subset.iterrows()
                        ]
                        hundred_volumes = [
                            v for v in hundred_volumes if v is not None and v > 0
                        ]

                        if not hundred_volumes:
                            return None

                        hundred_avg = mean(hundred_volumes)

                        if hundred_avg <= 0:
                            return None

                        return (period_avg / hundred_avg) * 100

                    except Exception:
                        return None

                # Volume ratios according to specification
                t_2_volyymi = calc_volume_ratio_spec(
                    2, -2
                )  # mean(t-2, t-1) / 100-day avg
                t_5_volyymi = calc_volume_ratio_spec(
                    5, -5
                )  # mean(t-5...t-1) / 100-day avg
                t_10_volyymi = calc_volume_ratio_spec(
                    10, -10
                )  # mean(t-10...t-1) / 100-day avg
                t_15_volyymi = calc_volume_ratio_spec(
                    15, -15
                )  # mean(t-15...t-1) / 100-day avg
                t_20_volyymi = calc_volume_ratio_spec(
                    20, -20
                )  # mean(t-20...t-1) / 100-day avg

                # t0 volume ratio (single day)
                t0_volume = safe_get(r0, vcol)
                hundred_start = max(0, idx - 100)
                hundred_end = idx - 1  # End at t-1
                if hundred_end >= hundred_start:
                    hundred_subset = df.iloc[hundred_start : hundred_end + 1]
                    hundred_volumes = [
                        safe_get(row, vcol) for _, row in hundred_subset.iterrows()
                    ]
                    hundred_volumes = [
                        v for v in hundred_volumes if v is not None and v > 0
                    ]
                    hundred_avg = mean(hundred_volumes) if hundred_volumes else None
                    t0_volyymi = (
                        (t0_volume / hundred_avg) * 100
                        if t0_volume and hundred_avg and hundred_avg > 0
                        else None
                    )
                else:
                    t0_volyymi = None

                # Future volume ratios
                t2_volyymi = calc_volume_ratio_spec(
                    2, 1
                )  # mean(t+1, t+2) / 100-day avg
                t5_volyymi = calc_volume_ratio_spec(
                    5, 1
                )  # mean(t+1...t+5) / 100-day avg
                t10_volyymi = calc_volume_ratio_spec(
                    10, 1
                )  # mean(t+1...t+10) / 100-day avg
                t20_volyymi = calc_volume_ratio_spec(
                    20, 1
                )  # mean(t+1...t+20) / 100-day avg

                # Calculate moving averages according to specification

                def calc_ma_normalized_spec(days_offset, ma_period):
                    """
                    Laskee liukuvan keskiarvon t0-päivää edeltävältä ajanjaksolta:
                    Esim. t_2_5p_liukuva = mean([t-2, t-3, t-4, t-5, t-6])
                    """
                    # Lasketaan päätepiste (t0 + days_offset), mutta EI oteta sitä mukaan, vaan taaksepäin ma_period päivää
                    end_idx = idx + days_offset  # t2 -> idx-2
                    start_idx = end_idx - ma_period + 1
                    if start_idx < 0 or end_idx < 0 or end_idx >= len(df):
                        return None
                    subset = df.iloc[start_idx : end_idx + 1]  # end_idx sisältyy
                    values = [safe_get(row, ccol) for _, row in subset.iterrows()]
                    values = [v for v in values if v is not None]
                    if len(values) != ma_period:
                        return None
                    ma_val = mean(values)
                    if is_index:
                        return (
                            (ma_val / t0_close * 100)
                            if t0_close and t0_close > 0
                            else None
                        )
                    else:
                        t0_low = safe_get(r0, lcol)
                        return (
                            (ma_val / t0_low * 100) if t0_low and t0_low > 0 else None
                        )

                t_2_5p_liukuva = calc_ma_normalized_spec(-2, 5)
                t_2_10p_liukuva = calc_ma_normalized_spec(-2, 10)
                t_2_20p_liukuva = calc_ma_normalized_spec(-2, 20)

                t_5_5p_liukuva = calc_ma_normalized_spec(-5, 5)
                t_5_10p_liukuva = calc_ma_normalized_spec(-5, 10)
                t_5_20p_liukuva = calc_ma_normalized_spec(-5, 20)

                t_10_5p_liukuva = calc_ma_normalized_spec(-10, 5)
                t_10_10p_liukuva = calc_ma_normalized_spec(-10, 10)
                t_10_20p_liukuva = calc_ma_normalized_spec(-10, 20)

                t_15_5p_liukuva = calc_ma_normalized_spec(-15, 5)
                t_15_10p_liukuva = calc_ma_normalized_spec(-15, 10)
                t_15_20p_liukuva = calc_ma_normalized_spec(-15, 20)

                t_20_5p_liukuva = calc_ma_normalized_spec(-20, 5)
                t_20_10p_liukuva = calc_ma_normalized_spec(-20, 10)
                t_20_20p_liukuva = calc_ma_normalized_spec(-20, 20)

                # t0_50p_liukuva: mean([t-49...t0]) / t0_low × 100
                t0_50p_liukuva = calc_ma_normalized_spec(0, 50)

                # Special case for t200_200p_liukuva - set to 0 if not enough data
                # Korjattu: etsitään t0:n indeksi päivämäärän perusteella
                t0_pvm = "2024-10-10"  # Testaa tällä päivällä, jatkossa parametrina
                if "pvm" in df.columns:
                    t0_idx = df.index[df["pvm"] == t0_pvm]
                    if len(t0_idx) > 0:
                        t0_idx = t0_idx[0]

                        def calc_ma_normalized_spec_idx(idx, days_offset, ma_period):
                            end_idx = idx + days_offset
                            start_idx = end_idx - ma_period + 1
                            if start_idx < 0 or end_idx < 0 or end_idx >= len(df):
                                return None
                            subset = df.iloc[start_idx : end_idx + 1]
                            values = [
                                safe_get(row, ccol) for _, row in subset.iterrows()
                            ]
                            values = [v for v in values if v is not None]
                            if len(values) != ma_period:
                                return None
                            ma_val = mean(values)
                            t0_low = safe_get(r0, lcol)
                            return (
                                (ma_val / t0_low * 100)
                                if t0_low and t0_low > 0
                                else None
                            )

                        t200_200p_liukuva = calc_ma_normalized_spec_idx(t0_idx, 0, 200)
                        if t200_200p_liukuva is None:
                            t200_200p_liukuva = 0
                    else:
                        t200_200p_liukuva = 0
                else:
                    t200_200p_liukuva = 0

                # Calculate index data according to specification
                def get_index_normalized_spec(index_ticker, offset, data_type="close"):
                    """
                    Get index data with specification normalization:
                    - Indices: normalized to t0_close=100 (not t0_low)
                    """
                    # Get the index's t0_close for normalization (not t0_low)
                    index_t0_close = _get_index_data(
                        oconn, index_ticker, date, 0, "close"
                    )
                    if index_t0_close is None or index_t0_close <= 0:
                        return None

                    index_value = _get_index_data(
                        oconn, index_ticker, date, offset, data_type
                    )
                    return (
                        (index_value / index_t0_close * 100)
                        if index_value is not None
                        else None
                    )

                # S&P 500 data (normalized to t0_close=100)
                SPX_0 = 100.0  # t0_close normalized to 100

                SPX_2 = get_index_normalized_spec("^GSPC", -2)
                SPX_5 = get_index_normalized_spec("^GSPC", -5)
                SPX_10 = get_index_normalized_spec("^GSPC", -10)
                SPX_15 = get_index_normalized_spec("^GSPC", -15)
                SPX_20 = get_index_normalized_spec("^GSPC", -20)

                SPX2 = get_index_normalized_spec("^GSPC", 2)
                SPX5 = get_index_normalized_spec("^GSPC", 5)
                SPX10 = get_index_normalized_spec("^GSPC", 10)
                SPX15 = get_index_normalized_spec("^GSPC", 15)
                SPX20 = get_index_normalized_spec("^GSPC", 20)

                # Nasdaq 100 data (normalized to t0_close=100)
                NDX_0 = 100.0  # t0_close normalized to 100

                NDX_2 = get_index_normalized_spec("^NDX", -2)
                NDX_5 = get_index_normalized_spec("^NDX", -5)
                NDX_10 = get_index_normalized_spec("^NDX", -10)
                NDX_15 = get_index_normalized_spec("^NDX", -15)
                NDX_20 = get_index_normalized_spec("^NDX", -20)

                NDX2 = get_index_normalized_spec("^NDX", 2)
                NDX5 = get_index_normalized_spec("^NDX", 5)
                NDX10 = get_index_normalized_spec("^NDX", 10)
                NDX15 = get_index_normalized_spec("^NDX", 15)
                NDX20 = get_index_normalized_spec("^NDX", 20)

                # RSI-14 t0-arvo (luettu analysis.db:stä)
                rsi14_t0 = None
                if rsi14_from_db is not None:
                    try:
                        rsi14_t0 = float(rsi14_from_db)
                    except Exception:
                        rsi14_t0 = None

                # Convert candle name to integer using mapping
                candle_int = CANDLE_MAPPING.get(candle, 0)  # 0 for unknown patterns

                # Format values for output (2 decimal places for Excel)
                def fmt_val(v, decimals=2):
                    if v is None:
                        return ""
                    if isinstance(v, (int, float)):
                        return round(v, decimals)
                    return v

                # Build output row with all 81 columns according to specification
                out = [
                    ticker,  # 1. osake
                    date,  # 2. date
                    candle_int,  # 3. kynttila (as integer)
                    (
                        fmt_val(signal_strength) if signal_strength is not None else ""
                    ),  # 4. vahvuus
                    # Detailed candle data
                    fmt_val(t_1_alin),  # 5. t_1_alin
                    fmt_val(t_1_ylin),  # 6. t_1_ylin
                    fmt_val(t_1_bodi),  # 7. t_1_bodi
                    fmt_val(t_1_bodi_colour),  # 8. t_1_bodi_colour
                    fmt_val(t0_alin),  # 9. t0_alin
                    fmt_val(t0_ylin),  # 10. t0_ylin
                    fmt_val(t0_bodi),  # 11. t0_bodi
                    fmt_val(t0_bodi_colour),  # 12. t0_bodi_colour
                    fmt_val(t1_alin),  # 13. t1_alin
                    fmt_val(t1_ylin),  # 14. t1_ylin
                    fmt_val(t1_bodi),  # 15. t1_bodi
                    fmt_val(t1_bodi_colour),  # 16. t1_bodi_colour
                    # Historical prices
                    fmt_val(t_2),  # 17. t_2
                    fmt_val(t_5),  # 18. t_5
                    fmt_val(t_10),  # 19. t_10
                    fmt_val(t_15),  # 20. t_15
                    fmt_val(t_20),  # 21. t_20
                    # Volatility (match header order 22-26)
                    fmt_val(t_2_hajonta),  # 22. t_2_hajonta
                    fmt_val(t_5_hajonta),  # 23. t_5_hajonta
                    fmt_val(t_10_hajonta),  # 24. t_10_hajonta
                    fmt_val(t_15_hajonta),  # 25. t_15_hajonta
                    fmt_val(t_20_hajonta),  # 26. t_20_hajonta
                    # Future prices (match header order 27-30)
                    fmt_val(t2),  # 27. t2
                    fmt_val(t5),  # 28. t5
                    fmt_val(t10),  # 29. t10
                    fmt_val(t20),  # 30. t20
                    # Volume ratios
                    fmt_val(t_2_volyymi),  # 31. t_2_volyymi
                    fmt_val(t_5_volyymi),
                    fmt_val(t_10_volyymi),
                    fmt_val(t_15_volyymi),
                    fmt_val(t_20_volyymi),
                    fmt_val(t0_volyymi),
                    fmt_val(t2_volyymi),
                    fmt_val(t5_volyymi),
                    fmt_val(t10_volyymi),
                    fmt_val(t20_volyymi),
                    # Moving averages
                    fmt_val(t_2_5p_liukuva),
                    fmt_val(t_2_10p_liukuva),
                    fmt_val(t_2_20p_liukuva),
                    fmt_val(t_5_5p_liukuva),
                    fmt_val(t_5_10p_liukuva),
                    fmt_val(t_5_20p_liukuva),
                    fmt_val(t_10_5p_liukuva),
                    fmt_val(t_10_10p_liukuva),
                    fmt_val(t_10_20p_liukuva),
                    fmt_val(t_15_5p_liukuva),
                    fmt_val(t_15_10p_liukuva),
                    fmt_val(t_15_20p_liukuva),
                    fmt_val(t_20_5p_liukuva),
                    fmt_val(t_20_10p_liukuva),
                    fmt_val(t_20_20p_liukuva),
                    fmt_val(t0_50p_liukuva),
                    fmt_val(t200_200p_liukuva),
                    # S&P 500 index
                    fmt_val(SPX_0),
                    fmt_val(SPX_2),
                    fmt_val(SPX_5),
                    fmt_val(SPX_10),
                    fmt_val(SPX_15),
                    fmt_val(SPX_20),
                    fmt_val(SPX2),
                    fmt_val(SPX5),
                    fmt_val(SPX10),
                    fmt_val(SPX15),
                    fmt_val(SPX20),
                    # Nasdaq 100 index
                    fmt_val(NDX_0),
                    fmt_val(NDX_2),
                    fmt_val(NDX_5),
                    fmt_val(NDX_10),
                    fmt_val(NDX_15),
                    fmt_val(NDX_20),
                    fmt_val(NDX2),
                    fmt_val(NDX5),
                    fmt_val(NDX10),
                    fmt_val(NDX15),
                    fmt_val(NDX20),
                    fmt_val(rsi14_t0),
                    fmt_val(t0_close_norm),
                    bearish_div,  # 82. Bearish Divergence vahvuus (0 tai 1.00-3.00)
                    bullish_div,  # 83. Bullish Divergence vahvuus (0 tai 1.00-3.00)
                ]

                output_rows.append(out)

    return header, output_rows


# ============================================================================
# Deprecated functions removed (moved to deprecated/DEPRECATED_FUNCTIONS.md):
# - paivita_results_csv() - old Excel generation with UI, replaced by excel_cache.py
# - paivita_results_csv_click() - event handler for old function
# ============================================================================


def generate_results_now(
    write: bool = True,
    downtrend_filter: bool = True,
    min_decline_percent: float = 3.0,
    use_ma_filter: bool = True,
    use_volume_filter: bool = False,
    progress_callback=None,
):
    """Generoi results.xlsx tiedosto

    Args:
        write: Kirjoitetaanko tiedostoon vai palautetaanko vain rivimäärä
        downtrend_filter: Suodatetaanko vain laskutrendien kynttilät
        min_decline_percent: Minimalasku prosentteina
        use_ma_filter: Käytetäänkö liukuva keskiarvo -suodatinta
        use_volume_filter: Käytetäänkö volyymi-suodatinta
        progress_callback: Callback-funktio edistymisen raportointiin
    """
    base = Path(__file__).resolve().parents[1]
    analysis_db = base / "data" / "analysis.db"
    osake_db = base / "data" / "osakedata.db"
    excel_path = base / "data" / "results.xlsx"

    if progress_callback:
        progress_callback("🔍 Analysoidaan tietoja...", 0.1)

    header, output_rows = _build_output_rows(
        analysis_db,
        osake_db,
        downtrend_filter,
        min_decline_percent,
        use_ma_filter,
        use_volume_filter,
        progress_callback,
    )

    if not output_rows:
        return 0

    added = len(output_rows)

    if write:
        if progress_callback:
            progress_callback("📊 Luodaan Excel-tiedostoa...", 0.8)

        # Luodaan Excel-tiedosto
        success = _create_excel_file(header, output_rows, excel_path, downtrend_filter)

        if success:
            print(f"✅ Excel-tiedosto luotu: {excel_path}")
        else:
            print("❌ Excel-tiedoston luonti epäonnistui")

        if progress_callback:
            progress_callback("✅ Valmis!", 1.0)

    return added


def generate_excel_optimized(
    excel_path: str = "data/results.xlsx",
    progress_callback=None,
    analysis_db: str = "data/analysis.db",
    osake_db: str = "data/osakedata.db",
    force_rebuild: bool = False,
    limit_rows: int = None,
    ticker_filter: str = None,
) -> int:
    """
    Optimoitu Excel-generointi staging-tietokannan avulla.

    Args:
        excel_path: Polku Excel-tiedostoon
        progress_callback: Funktio progress-päivityksille
        analysis_db: Polku analysis.db tietokantaan
        osake_db: Polku osakedata.db tietokantaan
        force_rebuild: Pakota staging-taulun uudelleenrakennus
        limit_rows: Rajoita rivien määrä (None = ei rajoitusta, testikäyttöön)
        ticker_filter: Rajoita tiettyyn tickeriin (None = kaikki tickerit)

    Returns:
        int: Löydösten määrä
    """

    try:
        # Luo cache-objekti
        cache = ExcelResultsCache(
            analysis_db_path=analysis_db,
            osake_db_path=osake_db,
            results_db_path="data/results.db",
        )

        # Tarkista onko cache tuore (mutta pakota rebuild jos ticker-filtteri käytössä)
        cache_is_fresh = not force_rebuild and cache.is_cache_fresh()
        print(
            f"🔍 Debug: force_rebuild={force_rebuild}, cache.is_cache_fresh()={cache.is_cache_fresh()}"
        )

        # Jos ticker-filtteri on käytössä, pakota rebuild
        if ticker_filter:
            print(
                f"🎯 Ticker-filtteri ({ticker_filter}) aktiivinen - pakotetaan rebuild"
            )
            cache_is_fresh = False

        print(f"🔍 Debug: cache_is_fresh={cache_is_fresh}")

        if cache_is_fresh:
            if progress_callback:
                progress_callback("📊 Käytetään cache-dataa...", 50, 100)

            # Nopea export staging-taulusta
            success = cache.export_to_excel_fast(
                excel_path, limit_rows=limit_rows, ticker_filter=ticker_filter
            )

            if success:
                stats = cache.get_staging_stats()
                if progress_callback:
                    progress_callback("✅ Valmis!", 100, 100)
                return stats.get("total_rows", 0)
            else:
                # Jos nopea export epäonnistui, rebuild cache
                force_rebuild = True

        if force_rebuild or not cache.is_cache_fresh():
            if progress_callback:
                progress_callback("🔄 Rakennetaan cache...", 10, 100)

            # Rebuild staging-taulu
            def cache_progress(step, current, total):
                # Skaalaa progress 10-90% välille
                scaled_progress = 10 + int((current / total) * 80)
                if progress_callback:
                    progress_callback(step, scaled_progress, 100)

            cache.rebuild_staging_optimized(
                cache_progress, limit_rows=limit_rows, ticker_filter=ticker_filter
            )

            if progress_callback:
                progress_callback("📊 Luodaan Excel-tiedostoa...", 90, 100)

            # Export Exceliin
            success = cache.export_to_excel_fast(
                excel_path, limit_rows=limit_rows, ticker_filter=ticker_filter
            )

            if success:
                stats = cache.get_staging_stats()
                if progress_callback:
                    progress_callback("✅ Valmis!", 100, 100)
                return stats.get("total_rows", 0)
            else:
                if progress_callback:
                    progress_callback("❌ Excel-generointi epäonnistui", 100, 100)
                return 0

        return 0

    except Exception as e:
        error_msg = f"❌ Virhe optimoidussa Excel-generoinnissa: {e}"
        print(error_msg)
        if progress_callback:
            progress_callback(error_msg, 100, 100)

        # Fallback: käytä vanhaa algoritmia
        print("🔄 Yritetään vanhaa algoritmia...")

        # Vanhan algoritmin progress callback on erilainen
        def old_progress_callback(message, progress):
            if progress_callback:
                # Muunna progress (0.0-1.0) -> (current, total) muotoon
                current = int(progress * 100)
                total = 100
                progress_callback(message, current, total)

        return generate_results_now(
            write=True,
            downtrend_filter=True,
            min_decline_percent=1.0,
            use_ma_filter=False,
            use_volume_filter=False,
            progress_callback=old_progress_callback,
        )
