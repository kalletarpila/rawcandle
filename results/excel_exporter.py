"""
Excel-vienti results_data taulusta.

Vie results_data taulun tiedot Excel-tiedostoon samassa muodossa
kuin alkuperäinen generate_results.py ja liittää uudet feature-sarakkeet
85. kolumnista eteenpäin.
"""

import logging
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from analysis.database_manager import DatabaseManager


class ExcelExporter:
    """Vie results_data Excel-tiedostoon."""

    # Pattern mappaus (numero -> nimi)
    PATTERN_NAMES = {
        0: "downtrend",
        1: "Hammer",
        2: "Bullish Engulfing",
        3: "Piercing Pattern",
        4: "Three White Soldiers",
        5: "Morning Star",
        6: "Dragonfly Doji",
        7: "Bullish Divergence",
        8: "Bearish Divergence",
        71: "BullDiv & Hammer",
        72: "BullDiv & Bullish Engulfing",
        73: "BullDiv & Piercing Pattern",
        74: "BullDiv & Three White Soldiers",
        75: "BullDiv & Morning Star",
        76: "BullDiv & Dragonfly Doji",
    }

    # Täysi otsikkolista: kaikki results_data-sarakkeet (id pois, päivämäärät ja markkina mukana)
    HEADERS = [
        "ticker",
        "date",
        "market",
        "candle_pattern",
        "signal_strength",
        "t_1_alin",
        "t_1_ylin",
        "t_1_bodi",
        "t_1_bodi_colour",
        "t0_alin",
        "t0_ylin",
        "t0_bodi",
        "t0_bodi_colour",
        "t0_alinMiinusClose",
        "t1_alin",
        "t1_ylin",
        "t1_bodi",
        "t1_bodi_colour",
        "t_2",
        "t_5",
        "t_10",
        "t_15",
        "t_20",
        "t_2_hajonta",
        "t_5_hajonta",
        "t_10_hajonta",
        "t_15_hajonta",
        "t_20_hajonta",
        "t2",
        "t5",
        "t10",
        "t20",
        "t_2_volyymi",
        "t_5_volyymi",
        "t_10_volyymi",
        "t_15_volyymi",
        "t_20_volyymi",
        "t0_volyymi",
        "t2_volyymi",
        "t5_volyymi",
        "t10_volyymi",
        "t20_volyymi",
        "t_2_5p_liukuva",
        "t_2_10p_liukuva",
        "t_2_20p_liukuva",
        "t_5_5p_liukuva",
        "t_5_10p_liukuva",
        "t_5_20p_liukuva",
        "t_10_5p_liukuva",
        "t_10_10p_liukuva",
        "t_10_20p_liukuva",
        "t_15_5p_liukuva",
        "t_15_10p_liukuva",
        "t_15_20p_liukuva",
        "t_20_5p_liukuva",
        "t_20_10p_liukuva",
        "t_20_20p_liukuva",
        "t0_20p_liukuva",
        "t0_50p_liukuva",
        "t0_200p_liukuva",
        "SPX_0",
        "SPX_2",
        "SPX_5",
        "SPX_10",
        "SPX_15",
        "SPX_20",
        "SPX2",
        "SPX5",
        "SPX10",
        "SPX15",
        "SPX20",
        "NDX_0",
        "NDX_2",
        "NDX_5",
        "NDX_10",
        "NDX_15",
        "NDX_20",
        "NDX2",
        "NDX5",
        "NDX10",
        "NDX15",
        "NDX20",
        "RSI14_t0",
        "t0_close_norm",
        "BullDiv_strength",
        "RSI_slope_5",
        "Price_slope_5",
        "Price_slope_10",
        "Price_acceleration_5_10",
        "Volatility_ratio_10_20",
        "Gap_down_strength",
        "Body_ratio",
        "Shadow_ratio",
        "SPX_volatility_10",
        "NDX_volatility_10",
        "Volume_impulse",
        "Reversal_Context_Score",
        "bullDiv_offset",
        # Trendit ja volatiliteetti / indikaattorit
        "t0_50p_slope",
        "t0_200p_slope",
        "trend_regime_5_20",
        "trend_regime_20_50",
        "trend_regime_50_200",
        "ATR_14",
        "ATR_ratio_14",
        "MACD_line",
        "MACD_signal",
        "MACD_hist",
        "VIX_10",
        "VIX_norm_10",
        "sector",
        "sector_momentum_5",
        "sector_momentum_20",
        "sector_volatility_20",
        "weekday",
    ]
    NON_ROUNDED_HEADERS = {
        "t_1_bodi_colour",
        "t0_bodi_colour",
        "t1_bodi_colour",
        "weekday",
        "BullDiv_recent_offset",
        "Has_BullDiv_recent",
        "bullDiv_offset",
        "bullDiv_last_1d",
        "bullDiv_last_2d",
        "bullDiv_last_3d",
        "bullDiv_last_3d_any",
        "is_candle_day",
        "is_crisis",
        "has_blackout_data",
        "is_earnings_t0",
        "is_earnings_window",
        "is_dividend_t0",
        "is_dividend_window",
        "is_blackout_t0",
        "is_blackout_window",
        "exclude_from_regression",
        "trend_regime_5_20",
        "trend_regime_20_50",
        "trend_regime_50_200",
    }
    # Kaikki REAL-luvut pyöristetään 10 desimaaliin, ellei NON_ROUNDED_HEADERS sisällä
    PRECISION_MAP = {  # arvot >0 sallii eri desimaalit mutta käytämme 10 oletuksena
        "BullDiv_recent_offset": 0,
        "Has_BullDiv_recent": 0,
        "bullDiv_offset": 0,
        "bullDiv_last_1d": 0,
        "bullDiv_last_2d": 0,
        "bullDiv_last_3d": 0,
        "bullDiv_last_3d_any": 0,
        "is_candle_day": 0,
        "is_crisis": 0,
        "has_blackout_data": 0,
        "is_earnings_t0": 0,
        "is_earnings_window": 0,
        "is_dividend_t0": 0,
        "is_dividend_window": 0,
        "is_blackout_t0": 0,
        "is_blackout_window": 0,
        "exclude_from_regression": 0,
        "trend_regime_5_20": 0,
        "trend_regime_20_50": 0,
        "trend_regime_50_200": 0,
    }

    def __init__(self, db_path: str = "analysis.db"):
        """
        Alusta exporter.

        Args:
            db_path: Polku analysis.db tietokantaan
        """
        self.db_path = db_path
        self.db_manager = DatabaseManager(db_path)
        self.logger = logging.getLogger(__name__)

    def _normalize_rows(self, rows):
        """Palauta listamuotoinen kopio annetusta rivilähteestä."""
        if rows is None:
            return []
        if isinstance(rows, list):
            return rows
        try:
            return list(rows)
        except TypeError:
            return []

    def _fetch_rows_by_ids(self, id_filter: list[int]) -> list[dict]:
        """Hae rivit annetuilla ID:illä, fallback-käsittelyllä testejä varten."""
        rows = self._normalize_rows(self.db_manager.get_results_by_ids(id_filter))
        if rows:
            return rows

        fallback = self._normalize_rows(self.db_manager.get_results_data())
        if not fallback:
            return []

        id_set = set(id_filter)
        return [row for row in fallback if row.get("id") in id_set]

    def _precision_for(self, header: str) -> int:
        if header in self.PRECISION_MAP:
            return self.PRECISION_MAP[header]
        return 10

    def export_to_excel(
        self,
        output_path: str,
        selected_patterns: Optional[list] = None,
        ticker_filter: Optional[list] = None,
        id_filter: Optional[list] = None,
        progress_callback=None,
        start_date: Optional[str] = None,
        end_date: Optional[str] = None,
        downtrend_only: bool = False,
        growth_limit: Optional[float] = None,
        drop_limit: Optional[float] = None,
    ) -> tuple[bool, str]:
        """
        Vie results_data Excel-tiedostoon.

        Args:
            output_path: Polku luotavaan Excel-tiedostoon
            selected_patterns: Lista pattern-numeroita joita viedään (None = kaikki)
            ticker_filter: Lista tickereistä joita viedään (None = kaikki)
            id_filter: Lista result ID:istä joita viedään (None = kaikki)
            progress_callback: Callback(current, total) -> bool (True = keskeytä)
            start_date: ISO-päivä (YYYY-MM-DD) alarajana (None = ei rajaa)
            end_date: ISO-päivä (YYYY-MM-DD) ylärajana (None = ei rajaa)
            downtrend_only: Jos True, vie vain downtrend-mallin rivit
            growth_limit: Maksimi sallittu arvo t2/t5/t10/t20:lle (None = ei rajaa)
            drop_limit: Minimi sallittu arvo t2/t5/t10/t20:lle (None = ei rajaa)

        Returns:
            Tuple[bool, str]: (success, message)
        """
        try:
            effective_patterns = None
            if selected_patterns is not None:
                effective_patterns = [p for p in selected_patterns if p != 8]
                if not effective_patterns:
                    return False, "Bearish Divergence ei ole Excel-viennissä käytössä."

            if id_filter is not None:
                results = self._fetch_rows_by_ids(id_filter)
            else:
                results = self._normalize_rows(
                    self.db_manager.get_results_filtered(
                        patterns=effective_patterns,
                        tickers=ticker_filter,
                        start_date=start_date,
                        end_date=end_date,
                        downtrend_only=downtrend_only,
                    )
                )

            original_total = len(results)
            results = [row for row in results if row.get("candle_pattern") != 8]
            removed_disabled = original_total - len(results)
            if not results:
                return False, "Valituilla suodattimilla ei löytynyt tuloksia."

            if len(results) != original_total:
                self.logger.info(
                    "Filtered out %s Bearish Divergence rows before Excel export",
                    original_total - len(results),
                )

            # Ääriarvofiltterit: drop None ja yli/ali rajat t2/t5/t10/t20-kentistä
            def _passes_extreme_filters(row: dict) -> bool:
                fields = ["t2", "t5", "t10", "t20"]
                values = []
                for f in fields:
                    val = row.get(f)
                    if val is None:
                        return False  # pudota None
                    try:
                        values.append(float(val))
                    except (TypeError, ValueError):
                        return False
                if growth_limit is not None and any(v > growth_limit for v in values):
                    return False
                if drop_limit is not None and any(v < drop_limit for v in values):
                    return False
                return True

            before_extremes = len(results)
            results = [row for row in results if _passes_extreme_filters(row)]
            removed_extremes = before_extremes - len(results)

            if not results:
                return False, "Valituilla suodattimilla ei löytynyt tuloksia."

            self.logger.info(f"Exporting {len(results)} results to Excel")
            if removed_extremes:
                self.logger.info(
                    "Filtered out %s rows by growth/drop limits (None also dropped)",
                    removed_extremes,
                )

            # Luo Excel-tiedosto
            wb = Workbook()
            ws = wb.active
            ws.title = "Kynttilätulokset"

            # Otsikkorivi (perussarakkeet + uudet featuret)
            headers = self.HEADERS

            # Tyylitys otsikoille
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )

            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill

            # Lisää data
            total_rows = len(results)
            skipped_rows = 0
            written_rows = 0

            for row_num, result in enumerate(results, 2):
                # Tarkista keskeytys joka 100. rivi
                if progress_callback and (row_num - 2) % 100 == 0:
                    cancelled = progress_callback(row_num - 2, total_rows)
                    if cancelled:
                        self.logger.info(
                            f"Excel export cancelled at row {row_num - 2}/{total_rows}"
                        )
                        return False, f"Vienti keskeytetty ({row_num - 2} riviä viety)"

                # Käytä pattern numeroa (SPSS-yhteensopivuus)
                pattern_num = result.get("candle_pattern", 0)

                bull_strength = result.get("BullDiv_strength")
                bull_recent_strength = result.get("BullDiv_recent_strength")
                bull_recent_offset = result.get("BullDiv_recent_offset")
                has_bull_recent = result.get("Has_BullDiv_recent")

                # Taaksepäinyhteensopivat oletusarvot
                if bull_strength is None:
                    bull_strength = 0.0
                if bull_recent_strength is None:
                    bull_recent_strength = bull_strength
                if bull_recent_offset is None:
                    bull_recent_offset = -1
                if has_bull_recent is None:
                    has_bull_recent = 1 if bull_recent_strength > 0 else 0

                # Rakenna rivi HEADERS-järjestyksessä
                row_data = []
                for header in self.HEADERS:
                    if header == "kynttila":
                        row_data.append(pattern_num)
                    elif header == "vahvuus":
                        row_data.append(result.get("signal_strength"))
                    elif header == "BullDiv_strength":
                        row_data.append(bull_strength)
                    elif header == "BullDiv_recent_strength":
                        row_data.append(bull_recent_strength)
                    elif header == "BullDiv_recent_offset":
                        row_data.append(bull_recent_offset)
                    elif header == "Has_BullDiv_recent":
                        row_data.append(has_bull_recent)
                    else:
                        row_data.append(result.get(header))

                # Kirjoita rivi (käytä written_rows laskuria rivin numeroinnissa)
                written_rows += 1
                actual_row = written_rows + 1  # +1 koska otsikkorivi on rivi 1

                for col_num, (header, value) in enumerate(zip(headers, row_data), 1):
                    # Pyöristä REAL-luvut 2 desimaaliin
                    if isinstance(value, (int, float)) and value is not None:
                        # Älä pyöristä kokonaislukuja (colour, weekday)
                        if header in self.NON_ROUNDED_HEADERS:
                            ws.cell(row=actual_row, column=col_num, value=value)
                        else:
                            precision = self._precision_for(header)
                            ws.cell(
                                row=actual_row,
                                column=col_num,
                                value=round(value, precision),
                            )
                    else:
                        ws.cell(row=actual_row, column=col_num, value=value)

            # Automaattinen sarakeleveys
            for col_num in range(1, len(headers) + 1):
                column_letter = get_column_letter(col_num)
                max_length = 0

                for cell in ws[column_letter]:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass

                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Tallenna
            wb.save(output_path)

            # Luo viesti
            message = f"Viety {written_rows} riviä Exceliin"
            if removed_disabled > 0:
                message += (
                    f" (pyydetty {original_total} riviä, {removed_disabled} "
                    "Bearish Divergence -riviä ohitettu)"
                )
            if skipped_rows > 0:
                message += (
                    f" ({skipped_rows} riviä ohitettu puuttuvien tietojen vuoksi)"
                )

            self.logger.info(f"Excel file saved: {output_path} - {message}")
            return True, message

        except Exception as e:
            self.logger.error(f"Export to Excel failed: {e}")
            return False, f"Excel-vienti epäonnistui: {e}"


if __name__ == "__main__":
    # Testi
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    )

    exporter = ExcelExporter("analysis.db")
    success, msg = exporter.export_to_excel("test_results.xlsx")
    print(f"Success: {success}")
    print(f"Message: {msg}")
