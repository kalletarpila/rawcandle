"""
Batch Simulator - Suorittaa simulaation kaikille osakkeille kannassa
"""

import os
import sqlite3
from datetime import datetime
from typing import Dict, List, Any, Callable
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill


class BatchSimulator:
    """Suorittaa simulaation kaikille osakkeille ja tallentaa tulokset Exceliin"""

    def __init__(self, data_dir: str = "data"):
        """
        Args:
            data_dir: Polku data-hakemistoon
        """
        self.data_dir = data_dir
        self.cancelled = False

    def get_all_tickers(self) -> List[str]:
        """
        Hae kaikki tickerit osakedata-kannasta

        Returns:
            Lista tickereistä
        """
        db_path = os.path.join(self.data_dir, "osakedata.db")

        if not os.path.exists(db_path):
            return []

        try:
            with sqlite3.connect(db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT DISTINCT osake FROM osakedata ORDER BY osake")
                return [row[0] for row in cursor.fetchall()]
        except Exception as e:
            print(f"Virhe tickerien haussa: {e}")
            return []

    def run_batch_simulation(
        self,
        simulation_func: Callable,
        parameters: Dict[str, Any],
        progress_callback: Callable[[int, int, str], None] = None,
    ) -> str:
        """
        Suorita simulaatio kaikille osakkeille

        Args:
            simulation_func: Funktio joka suorittaa simulaation yhdelle osakkeelle
                            Palauttaa: (success: bool, results: dict, error: str)
            parameters: Simulaation parametrit
            progress_callback: Callback edistymisen raportointiin (current, total, ticker)

        Returns:
            Polku luotuun Excel-tiedostoon
        """
        self.cancelled = False
        tickers = self.get_all_tickers()

        if not tickers:
            raise ValueError("Ei osakkeita kannassa!")

        total = len(tickers)
        results = []

        # Suorita simulaatio jokaiselle osakkeelle
        for idx, ticker in enumerate(tickers, 1):
            if self.cancelled:
                break

            if progress_callback:
                progress_callback(idx, total, ticker)

            try:
                success, sim_results, error = simulation_func(ticker, parameters)

                if success:
                    # Lisää vain jos kauppoja on tehty
                    total_trades = sim_results.get("total_trades", 0)
                    if total_trades > 0:
                        results.append(
                            {
                                "ticker": ticker,
                                "status": "OK",
                                "initial_capital": sim_results.get(
                                    "initial_capital", 0
                                ),
                                "final_capital": sim_results.get("final_capital", 0),
                                "profit_eur": sim_results.get("profit_eur", 0),
                                "profit_pct": sim_results.get("profit_pct", 0),
                                "total_trades": total_trades,
                                "winning_trades": sim_results.get("winning_trades", 0),
                                "losing_trades": sim_results.get("losing_trades", 0),
                                "error": "",
                            }
                        )
                else:
                    results.append(
                        {
                            "ticker": ticker,
                            "status": "ERROR",
                            "initial_capital": parameters.get("initial_capital", 0),
                            "final_capital": 0,
                            "profit_eur": 0,
                            "profit_pct": 0,
                            "total_trades": 0,
                            "winning_trades": 0,
                            "losing_trades": 0,
                            "error": error or "Tuntematon virhe",
                        }
                    )

            except Exception as ex:
                results.append(
                    {
                        "ticker": ticker,
                        "status": "ERROR",
                        "initial_capital": parameters.get("initial_capital", 0),
                        "final_capital": 0,
                        "profit_eur": 0,
                        "profit_pct": 0,
                        "total_trades": 0,
                        "winning_trades": 0,
                        "losing_trades": 0,
                        "error": str(ex),
                    }
                )

        # Lajittele tulokset kasvuprosentin mukaan suurimmasta pienimpään
        results.sort(key=lambda x: x["profit_pct"], reverse=True)

        # Tallenna Excel
        excel_path = self._save_to_excel(results, parameters)
        return excel_path

    def cancel(self):
        """Keskeytä batch-simulaatio"""
        self.cancelled = True

    def _save_to_excel(
        self, results: List[Dict[str, Any]], parameters: Dict[str, Any]
    ) -> str:
        """
        Tallenna tulokset Excel-tiedostoon

        Args:
            results: Simulaation tulokset
            parameters: Käytetyt parametrit

        Returns:
            Polku luotuun tiedostoon
        """
        # Etsi seuraava vapaa numero
        counter = 1
        while True:
            filename = f"financials_{counter}.xlsx"
            filepath = os.path.join(self.data_dir, filename)
            if not os.path.exists(filepath):
                break
            counter += 1

        # Luo workbook
        wb = openpyxl.Workbook()

        # Poista default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Luo Parametrit-välilehti
        self._create_parameters_sheet(wb, parameters)

        # Luo Tulokset-välilehti
        self._create_results_sheet(wb, results)

        # Tallenna
        wb.save(filepath)
        return filepath

    def _create_parameters_sheet(
        self, wb: openpyxl.Workbook, parameters: Dict[str, Any]
    ):
        """Luo Parametrit-välilehti"""
        ws = wb.create_sheet("Parametrit", 0)

        # Otsikko
        ws["A1"] = "Parametri"
        ws["B1"] = "Arvo"
        ws["A1"].font = Font(bold=True)
        ws["B1"].font = Font(bold=True)

        row = 2

        # Timestamp
        ws[f"A{row}"] = "Ajoaika"
        ws[f"B{row}"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        row += 1

        # Parametrit
        for key, value in parameters.items():
            ws[f"A{row}"] = key
            ws[f"B{row}"] = str(value)
            row += 1

        # Leveydet
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 50

    def _create_results_sheet(
        self, wb: openpyxl.Workbook, results: List[Dict[str, Any]]
    ):
        """Luo Tulokset-välilehti"""
        ws = wb.create_sheet("Tulokset", 1)

        # Otsikot
        headers = [
            "Ticker",
            "Status",
            "Alkupääoma",
            "Loppupääoma",
            "Voitto/Tappio (€)",
            "Kasvu (%)",
            "Kauppoja",
            "Voitolliset",
            "Tappiolliset",
            "Virhe",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")

        # Tulokset
        for row_idx, result in enumerate(results, 2):
            ws.cell(row=row_idx, column=1).value = result["ticker"]
            ws.cell(row=row_idx, column=2).value = result["status"]

            # Numerot todellisina numeroina (float), aseta numeroformaatti pilkulla
            cell = ws.cell(row=row_idx, column=3)
            cell.value = result["initial_capital"]
            cell.number_format = "#,##0.00"

            cell = ws.cell(row=row_idx, column=4)
            cell.value = result["final_capital"]
            cell.number_format = "#,##0.00"

            cell = ws.cell(row=row_idx, column=5)
            cell.value = result["profit_eur"]
            cell.number_format = "#,##0.00"

            cell = ws.cell(row=row_idx, column=6)
            cell.value = result["profit_pct"]
            cell.number_format = "#,##0.00"

            ws.cell(row=row_idx, column=7).value = result["total_trades"]
            ws.cell(row=row_idx, column=8).value = result["winning_trades"]
            ws.cell(row=row_idx, column=9).value = result["losing_trades"]
            ws.cell(row=row_idx, column=10).value = result["error"]

            # Väritä ERROR-rivit punaisella
            if result["status"] == "ERROR":
                for col in range(1, 11):
                    ws.cell(row=row_idx, column=col).fill = PatternFill(
                        start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
                    )

        # Leveydet
        ws.column_dimensions["A"].width = 12  # Ticker
        ws.column_dimensions["B"].width = 10  # Status
        ws.column_dimensions["C"].width = 15  # Alkupääoma
        ws.column_dimensions["D"].width = 15  # Loppupääoma
        ws.column_dimensions["E"].width = 18  # Voitto €
        ws.column_dimensions["F"].width = 15  # Kasvu %
        ws.column_dimensions["G"].width = 12  # Kauppoja
        ws.column_dimensions["H"].width = 15  # Voitolliset
        ws.column_dimensions["I"].width = 15  # Tappiolliset
        ws.column_dimensions["J"].width = 40  # Virhe

    def _format_number(self, value: float) -> str:
        """
        Formatoi numero: 2 desimaalia, pilkku desimaalimerkkinä, ei tuhaterotinta

        Args:
            value: Numero

        Returns:
            Formatoitu merkkijono
        """
        try:
            # 2 desimaalin tarkkuus
            formatted = f"{float(value):.2f}"
            # Vaihda piste pilkuksi
            formatted = formatted.replace(".", ",")
            return formatted
        except (ValueError, TypeError):
            return "0,00"
