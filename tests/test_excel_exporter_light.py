import sqlite3
from pathlib import Path

from openpyxl import load_workbook

from analysis.database_manager import DatabaseManager
from results.excel_exporter import ExcelExporter


def _prepare_results_db(tmp_path: Path, row_overrides=None) -> str:
    db_path = tmp_path / "analysis.db"
    manager = DatabaseManager(str(db_path))
    conn = manager.get_connection()
    cursor = conn.cursor()

    cursor.execute("PRAGMA table_info(results_data)")
    columns_info = cursor.fetchall()

    def build_row(overrides: dict | None = None) -> dict:
        values = {}
        for _cid, name, col_type, *_rest in columns_info:
            if name == "id":
                continue
            upper_type = (col_type or "").upper()
            if name == "ticker":
                values[name] = "AAA"
            elif name == "date":
                values[name] = "2024-01-02"
            elif name == "candle_pattern":
                values[name] = 1
            elif name == "signal_strength":
                values[name] = 0.8
            elif name == "weekday":
                values[name] = 2
            elif "TEXT" in upper_type:
                values[name] = "0"
            elif "INTEGER" in upper_type:
                values[name] = 1
            else:  # REAL tai muu numeerinen
                values[name] = 1.0
        if overrides:
            values.update(overrides)
        return values

    rows = row_overrides or [{}]
    for overrides in rows:
        row_values = build_row(overrides)
        columns = ", ".join(row_values.keys())
        placeholders = ", ".join(["?"] * len(row_values))
        cursor.execute(
            f"INSERT INTO results_data ({columns}) VALUES ({placeholders})",
            list(row_values.values()),
        )

    conn.commit()
    return str(db_path)


def test_excel_exporter_writes_rows(tmp_path):
    db_path = _prepare_results_db(Path(tmp_path))
    exporter = ExcelExporter(db_path)

    output_file = Path(tmp_path) / "results.xlsx"
    success, message = exporter.export_to_excel(str(output_file))

    assert success, message
    assert output_file.exists()

    wb = load_workbook(output_file)
    ws = wb.active
    assert ws["A2"].value == "AAA"
    assert ws["C2"].value == "0"
    assert ws["D2"].value == 1


def test_excel_exporter_applies_filters(tmp_path):
    overrides = [
        {"ticker": "AAA", "date": "2024-01-02", "candle_pattern": 1},
        {"ticker": "BBB", "date": "2024-02-10", "candle_pattern": 0},
        {"ticker": "BBB", "date": "2024-03-15", "candle_pattern": 1},
    ]
    db_path = _prepare_results_db(Path(tmp_path), overrides)
    exporter = ExcelExporter(db_path)

    output_file = Path(tmp_path) / "filtered.xlsx"
    success, message = exporter.export_to_excel(
        str(output_file),
        ticker_filter=["BBB"],
        start_date="2024-02-01",
        end_date="2024-02-28",
        downtrend_only=True,
    )

    assert success, message
    wb = load_workbook(output_file)
    ws = wb.active
    assert ws.max_row == 2  # header + 1 row
    assert ws["A2"].value == "BBB"
    assert ws["B2"].value == "2024-02-10"
    assert ws["D2"].value == 0


def test_excel_exporter_extreme_filters_drop_rows(tmp_path):
    overrides = [
        {"ticker": "AAA", "t2": 200, "t5": 120, "t10": 110, "t20": 105},
        {"ticker": "BBB", "t2": 140, "t5": 130, "t10": 125, "t20": 115},
        {"ticker": "CCC", "t2": None, "t5": 90, "t10": 95, "t20": 100},
    ]
    db_path = _prepare_results_db(Path(tmp_path), overrides)
    exporter = ExcelExporter(db_path)

    output_file = Path(tmp_path) / "extreme.xlsx"
    success, message = exporter.export_to_excel(
        str(output_file),
        growth_limit=150.0,
        drop_limit=80.0,
    )

    assert success, message
    wb = load_workbook(output_file)
    ws = wb.active
    # Vain BBB jää (ei ylitä 150, ei alita 80, eikä None-arvoja)
    assert ws.max_row == 2  # header + BBB
    assert ws["A2"].value == "BBB"


def test_excel_exporter_includes_market_column(tmp_path):
    db_path = _prepare_results_db(
        Path(tmp_path),
        [{"ticker": "AAA", "market": "usa", "candle_pattern": 1}],
    )
    exporter = ExcelExporter(db_path)

    output_file = Path(tmp_path) / "market.xlsx"
    success, message = exporter.export_to_excel(str(output_file))

    assert success, message
    wb = load_workbook(output_file)
    ws = wb.active
    headers = [ws.cell(row=1, column=idx).value for idx in range(1, 6)]
    assert headers[:4] == ["ticker", "date", "market", "candle_pattern"]
    assert ws["C2"].value == "usa"
