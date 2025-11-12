"""
Integration tests that exercise the full analysis → results → Excel pipeline.

These tests build temporary SQLite databases for both the analysis data and
stock price history so that we can run the same modules the UI uses without
touching production files.
"""

from __future__ import annotations

import sqlite3
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, Iterable, Tuple

from openpyxl import load_workbook

from analysis.database_manager import DatabaseManager
from analysis.results_generator import ResultsGenerator
from analysis.run_analysis import run_candlestick_analysis
from results.excel_exporter import ExcelExporter

BASE_START = datetime(2024, 1, 1)
TOTAL_DAYS = 260


def _create_stock_database(
    db_path: Path,
    tickers: Iterable[str],
    overrides: Dict[Tuple[str, int], Dict[str, float]] | None = None,
) -> None:
    """Create osakedata.db with deterministic prices for all requested tickers."""
    overrides = overrides or {}
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute(
        """
        CREATE TABLE osakedata (
            osake TEXT,
            pvm TEXT,
            open REAL,
            high REAL,
            low REAL,
            close REAL,
            volume INTEGER,
            market TEXT NOT NULL DEFAULT 'usa'
        )
        """
    )

    def _insert_series(symbol: str, base_price: float, slope: float) -> None:
        for day in range(TOTAL_DAYS):
            date = (BASE_START + timedelta(days=day)).strftime("%Y-%m-%d")
            close = base_price + day * slope
            open_price = close - 0.4
            high = close + 0.6
            low = close - 0.6
            volume = 1_000_000 + day * 1_000

            override = overrides.get((symbol, day))
            if override:
                open_price = override.get("open", open_price)
                high = override.get("high", high)
                low = override.get("low", low)
                close = override.get("close", close)
                volume = override.get("volume", volume)

            cursor.execute(
                """
                INSERT INTO osakedata (osake, pvm, open, high, low, close, volume)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (symbol, date, open_price, high, low, close, volume),
            )

    for ticker in tickers:
        base = 120 + (abs(hash(ticker)) % 25)
        _insert_series(ticker, base, 0.5)

    # Indices used for normalization inside ResultsGenerator
    _insert_series("^GSPC", 4300.0, 5.0)
    _insert_series("^NDX", 15000.0, 10.0)

    conn.commit()
    conn.close()


def _seed_divergence_data(
    db_manager: DatabaseManager,
    ticker: str,
    bullish_dates: Iterable[str] | None = None,
    bearish_dates: Iterable[str] | None = None,
) -> None:
    """Populate divergence_data for all dates with optional highlights."""
    bullish_dates = set(bullish_dates or [])
    bearish_dates = set(bearish_dates or [])
    records = []

    for day in range(TOTAL_DAYS):
        date = (BASE_START + timedelta(days=day)).strftime("%Y-%m-%d")
        bullish = 0.7 if date in bullish_dates else 0.0
        bearish = 0.6 if date in bearish_dates else 0.0
        records.append((date, bullish, bearish, 55.0))

    if records:
        db_manager.save_divergence_batch(ticker, records)


def _run_analysis_for_ticker(
    db_manager: DatabaseManager, stock_db: Path, ticker: str
) -> Iterable[str]:
    """Run the candlestick analysis and persist findings to analysis.db."""
    analysis_results = run_candlestick_analysis(
        str(stock_db),
        ticker=ticker,
        patterns=["Hammer"],
        downtrend_filter=False,
    )
    assert analysis_results, "Expected analyzer to find at least one pattern"
    saved_dates = set()

    for key, entries in analysis_results.items():
        symbol, date = key.split("|", 1)
        for entry in entries:
            db_manager.save_finding(
                ticker=symbol,
                date=date,
                pattern=entry["pattern"],
                signal_strength=entry["strength"],
                rsi14=entry.get("rsi14"),
            )
            saved_dates.add(date)

    return saved_dates


def test_full_pipeline_generates_excel(tmp_path):
    """End-to-end integration: analysis -> results_data -> Excel export."""
    analysis_db = Path(tmp_path) / "analysis.db"
    stock_db = Path(tmp_path) / "stock.db"
    hammer_index = 220
    overrides = {
        ("TST1", hammer_index): {
            "open": 110.0,
            "close": 111.0,
            "low": 100.0,
            "high": 111.3,
        }
    }
    _create_stock_database(stock_db, ["TST1"], overrides=overrides)

    db_manager = DatabaseManager(str(analysis_db))
    found_dates = _run_analysis_for_ticker(db_manager, str(stock_db), "TST1")
    _seed_divergence_data(db_manager, "TST1", bullish_dates=found_dates)

    generator = ResultsGenerator(db_manager, str(stock_db))
    rows_inserted, _ = generator.generate_results()
    results = db_manager.get_results_data()

    assert rows_inserted > 0
    assert len(results) == 1
    assert results[0]["ticker"] == "TST1"

    exporter = ExcelExporter(str(analysis_db))
    output_file = Path(tmp_path) / "results.xlsx"
    success, message = exporter.export_to_excel(str(output_file))
    assert success, message
    assert output_file.exists()

    workbook = load_workbook(output_file)
    ws = workbook.active
    assert ws.max_row == len(results) + 1  # header + data rows
    assert ws["A2"].value == "TST1"
    assert ws["B2"].value in found_dates

    db_manager.close()
    exporter.db_manager.close()


def test_divergence_combo_filter_requires_matching_patterns(tmp_path):
    """Integration test for divergence combo filter end-to-end."""
    analysis_db = Path(tmp_path) / "analysis.db"
    stock_db = Path(tmp_path) / "stock.db"
    _create_stock_database(stock_db, ["COMBO", "SOLO"])

    db_manager = DatabaseManager(str(analysis_db))
    combo_date = (BASE_START + timedelta(days=30)).strftime("%Y-%m-%d")
    solo_date = (BASE_START + timedelta(days=35)).strftime("%Y-%m-%d")

    db_manager.save_finding("COMBO", combo_date, "Hammer", 0.91, 54.0)
    db_manager.save_finding("COMBO", combo_date, "Bullish Divergence", 0.6, 54.0)
    db_manager.save_finding("SOLO", solo_date, "Hammer", 0.88, 53.0)

    _seed_divergence_data(db_manager, "COMBO", bullish_dates=[combo_date])

    generator = ResultsGenerator(db_manager, str(stock_db))
    rows_inserted, _ = generator.generate_results(divergence_combo_filter=True)
    results = db_manager.get_results_data()

    assert rows_inserted > 0
    assert len(results) == 1
    assert results[0]["ticker"] == "COMBO"
    assert results[0]["date"] == combo_date

    db_manager.close()
